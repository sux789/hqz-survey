# -*- coding: utf-8 -*-
"""xlsx 导出器 — 双模板导出（基本信息 + 样地）+ 轨迹 GPX。

设计原则（2026-08 重构，不兼容旧数据）：
  - 仅三个分类：人工造林(table1)/封山育林(table2)/退化林修复(table3)
  - 基本信息：tpl/tpl-base.xlsx（2023年度官方模板，sheet=2023年度人工造林/封山育林/退化林，预填示例）；导出前清空数据区再填值
  - 样地：tpl/tpl-samples.xlsx 块结构（39 行/块）；每分类一个 sheet，每小班一块向下复制；导出前清空块1示例
  - 样地统计（苗木合格率组，表1 AN/AO/AP；表2 BE/BF/BG；表3 AS/AT/AU）：
      小班查数株数 = 调查总株数（与样地模板 B34 同口径）
                     = round(Σ种植÷个数÷150×单个网格面积×种植网格数量)，
                     个数 = 手写 sm_total_count（>0 生效）回退实际样地数
      合格率      = Σ成活÷Σ种植（比率 0-1，如 0.9524；模板列已设 0.00%
                     百分比格式，Excel 显示 95.24%；分派公式阈值 0.9/0.4
                     亦按比率比较——禁止 ×100 落库/落表）
      合格株树    = round(查数株数×合格率)
  - 手写签字导出为图片：签字 canvas 存 data_json（inspector_sign /
    co_inspector_sign，PNG data URL），导出时裁白边+等比缩放后以图片
    插入签字列（表1 AR/AU、表2 BI/BL、表3 AV/AY）
  - 样地页汇总手写项（sm_* 键）写入样地模板 B27-B39（个数/网格面积/数量/
    撑杆/覆膜/验收人/验收日期/备注）；样地行备注（sample.remark）写 I 列
  - 死亡株数 = 样地模板 E 列公式（种植-成活）自动算，不录入
  - 数据行按（林班, 调查小班号）数字序
"""
import base64
import io
import re
from copy import copy as _style_copy
from datetime import date as _date
from pathlib import Path

import openpyxl

from survey.core import schema as S
from survey.core import storage
from survey.core import gdb as GDB

# 模板文件路径（统一项目根 tpl/ 目录，固定文件名——换官方新模板时
# 直接替换同名文件即可，无需改代码；官方原版归档在 tpl/official/ 仅供对照）
# 基本信息模板：tpl-base.xlsx（3 个分类 sheet：2023年度人工造林 /
#   2023年度封山育林 / 2023年度退化林，官方预填示例），
#   导出前必须先清空数据区（见 _clear_*），deploy.sh 会同步 tpl/。
# 样地模板：tpl-samples.xlsx 清理版（块结构已清示例），
#   布局（2026-08-21 起去掉项目类型行、加验收人/验收日期，39 行一块）：
#   R1 标题（含调查小班号）/ R2 年度县乡+坐标+照片 / R3 列头 /
#   R4-26 数据槽（23 个）/ R27-39 汇总区：
#   B27 总样地个数(代码写) / B28-30 SUM公式 / B31 成活率公式(除0守卫) /
#   B32 单个网格面积 / B33 种植网格数量(手写录入) /
#   B34 调查总株数公式(除0守卫) / B35 撑杆 / B36 覆膜 / B37 验收人 /
#   B38 验收日期 / B39 备注（均手写录入，data_json.sm_* 键）。
_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
_TPL_DIR = _PROJECT_ROOT / "tpl"
_BASE_TEMPLATE = _TPL_DIR / "tpl-base.xlsx"
_SAMPLE_TEMPLATE = _TPL_DIR / "tpl-samples.xlsx"


def check_templates():
    """导出模板自检：文件存在 + 关键结构符合预期。

    模板缺失或被改动结构（sheet 改名/删列）时抛 RuntimeError 并给出精确
    原因与修复提示——避免导出时才以 FileNotFoundError/KeyError 等隐晦
    错误失败。launcher 启动时调用一次（醒目告警），导出前再调用兜底。
    """
    problems = []
    if not _BASE_TEMPLATE.exists():
        problems.append(
            f"缺少基本信息模板: {_BASE_TEMPLATE}"
            "（tpl/ 由 deploy.sh 同步；本地请确认 git 仓库完整）"
        )
    else:
        wb = openpyxl.load_workbook(_BASE_TEMPLATE, read_only=True)
        names = wb.sheetnames
        wb.close()
        missing = [s for s in _TPL_SHEET_NAMES.values() if s not in names]
        if missing:
            problems.append(
                f"{_BASE_TEMPLATE.name} 缺少分类 sheet {missing}（现有 {names}）"
                "——模板被改动，需同步修改 exporter._TPL_SHEET_NAMES"
            )
    if not _SAMPLE_TEMPLATE.exists():
        problems.append(f"缺少样地模板: {_SAMPLE_TEMPLATE}")
    if problems:
        raise RuntimeError("导出模板自检失败：\n" + "\n".join(f"  - {p}" for p in problems))

# ── 基本信息模板列 → schema key 映射（基于 tpl-base 第3-4行表头，2026-08-20 新布局）──
# value = (schema_key, source)
# source: prefilled 小班预填（GDB 绿色列） | input 调查录入
#         extra 扩展数据(轨迹/照片) | sample_stat 样地聚合统计
#         sign 手写签字（data_json 里的 PNG data URL → 以图片插入单元格）：
#   表1 AR/AU、表2 BI/BL、表3 AV/AY。
# 打卡坐标列统一走 input（录入值优先，回退 extras 打卡记录）。
_TPL_COL_MAPS = {
    "table1": {  # 表1－人工造林小班验收因子表
        'A': ('city', 'prefilled'),
        'B': ('subcompartment', 'prefilled'),       # 调查小班号
        'C': ('county', 'prefilled'),
        'D': ('township', 'prefilled'),
        'E': ('village', 'prefilled'),
        'F': ('forest_compartment', 'prefilled'),
        'G': ('subcompartment_orig', 'prefilled'),  # 小班（GDB 原值，可含汉字；无则回退调查小班号）
        'H': ('check_type', 'prefilled'),
        'I': ('project_name', 'prefilled'),
        'J': ('plan_year', 'prefilled'),
        'K': ('work_year', 'prefilled'),
        'L': ('ownership', 'prefilled'),
        'M': ('tree_species', 'prefilled'),
        'N': ('reported_area', 'prefilled'),
        'O': ('survival_pass', 'input'),
        'P': ('survival_replant', 'input'),
        'Q': ('survival_fail', 'input'),
        'R': ('verified_total', 'input'),
        'S': ('verified_pass', 'input'),
        'T': ('verified_replant', 'input'),
        'U': ('verified_fail', 'input'),
        'V': ('verified_loss', 'input'),
        'W': ('area_short_reason', 'input'),
        'X': ('unqualified_reason', 'input'),
        'Y': ('loss_reason', 'input'),
        'Z': ('mgmt_design', 'input'),
        'AA': ('mgmt_meeting', 'input'),
        'AB': ('mgmt_speech', 'input'),
        'AC': ('mgmt_survey', 'input'),
        'AD': ('mgmt_supervision', 'input'),
        'AE': ('mgmt_photo', 'input'),
        'AF': ('construction_area', 'input'),
        'AG': ('construction_rate', 'input'),
        'AH': ('archive_area', 'input'),
        'AI': ('archive_rate', 'input'),
        'AJ': ('protect_area', 'input'),
        'AK': ('protect_rate', 'input'),
        'AL': ('tend_area', 'input'),
        'AM': ('tend_rate', 'input'),
        'AN': ('planted_total', 'sample_stat'),     # 小班查数株数 = 调查总株数（B34 口径）
        'AO': ('qualified_count', 'sample_stat'),   # 合格株树 = 查数株数×合格率
        'AP': ('qualified_rate', 'sample_stat'),    # 合格率 = Σ成活/Σ种植（比率，0.00% 格式显示）
        'AQ': ('design_count', 'prefilled'),        # 小班设计株树（GDB 小班设计株树/需苗量）
        'AR': ('inspector_sign', 'sign'),           # 验收人员手写签字（PNG 图片）
        'AS': ('inspect_time', 'input'),
        'AT': ('remark', 'input'),
        'AV': ('sample_coord_x', 'input'),          # 打卡坐标x（回退 extras 打卡）
        'AW': ('sample_coord_y', 'input'),          # 打卡坐标y
        'AX': ('track', 'extra'),                   # 轨迹 GPX 文件名
        'AY': ('photos', 'extra'),                  # 小班照片文件名
        'AU': ('co_inspector_sign', 'sign'),        # 配合验收人员手写签字（PNG 图片）
    },
    "table2": {  # 表2－封山育林验收因子表（新布局：B调查小班号 F小班 R造林树种 T补植面积 AF优势树种 AG封前地类 AI郁闭度 BH小班设计株树）
        'A': ('city', 'prefilled'),
        'B': ('subcompartment', 'prefilled'),       # 调查小班号
        'C': ('county', 'prefilled'),
        'D': ('township', 'prefilled'),
        'E': ('village', 'prefilled'),
        'F': ('subcompartment_orig', 'prefilled'),  # 小班（GDB 原值，可含汉字；无则回退调查小班号）
        'G': ('check_type', 'prefilled'),
        'H': ('project_name', 'prefilled'),
        'I': ('plan_year', 'prefilled'),
        'J': ('start_year', 'prefilled'),
        'K': ('ownership', 'prefilled'),
        'L': ('seal_target', 'prefilled'),
        'M': ('seal_years', 'prefilled'),
        'N': ('seal_type', 'prefilled'),
        'O': ('seal_method', 'prefilled'),
        'P': ('seal_measure', 'prefilled'),
        'Q': ('forest_measure', 'prefilled'),
        'R': ('tree_species', 'prefilled'),         # 造林树种（GDB 绿色）
        'S': ('reported_area', 'prefilled'),
        'T': ('replant_area', 'prefilled'),         # 补植面积（GDB 绿色）
        'U': ('survival_pass', 'input'),
        'V': ('survival_replant', 'input'),
        'W': ('survival_fail', 'input'),
        'X': ('verified_total', 'input'),
        'Y': ('verified_pass', 'input'),
        'Z': ('verified_replant', 'input'),
        'AA': ('verified_fail', 'input'),
        'AB': ('verified_loss', 'input'),
        'AC': ('area_short_reason', 'input'),
        'AD': ('unqualified_reason', 'input'),
        'AE': ('loss_reason', 'input'),
        'AF': ('dominant_species', 'prefilled'),    # 优势树种（GDB 绿色）
        'AG': ('pre_land_type', 'prefilled'),       # 封前地类（GDB 绿色）
        'AH': ('cur_land_type', 'input'),           # 现地类（现地调查）
        'AI': ('canopy_cover', 'prefilled'),        # 郁闭度（GDB 绿色）
        'AJ': ('conifer_mother', 'input'),
        'AK': ('conifer_seedling', 'input'),
        'AL': ('conifer_sapling', 'input'),
        'AM': ('broadleaf_mother', 'input'),
        'AN': ('broadleaf_seedling', 'input'),
        'AO': ('broadleaf_sapling', 'input'),
        'AP': ('bamboo_count', 'input'),
        'AQ': ('mgmt_design', 'input'),
        'AR': ('mgmt_meeting', 'input'),
        'AS': ('mgmt_speech', 'input'),
        'AT': ('mgmt_survey', 'input'),
        'AU': ('mgmt_supervision', 'input'),
        'AV': ('mgmt_photo', 'input'),
        'AW': ('construction_area', 'input'),
        'AX': ('construction_rate', 'input'),
        'AY': ('archive_area', 'input'),
        'AZ': ('archive_rate', 'input'),
        'BA': ('protect_area', 'input'),
        'BB': ('protect_rate', 'input'),
        'BC': ('tend_area', 'input'),
        'BD': ('tend_rate', 'input'),
        'BE': ('planted_total', 'sample_stat'),     # 小班查数株数 = 调查总株数（B34 口径）
        'BF': ('qualified_count', 'sample_stat'),   # 合格株树
        'BG': ('qualified_rate', 'sample_stat'),    # 合格率
        'BH': ('design_count', 'prefilled'),        # 小班设计株树（GDB 绿色）
        'BI': ('inspector_sign', 'sign'),           # 验收人员手写签字（PNG 图片）
        'BJ': ('inspect_time', 'input'),
        'BK': ('remark', 'input'),
        'BL': ('co_inspector_sign', 'sign'),        # 配合验收人员手写签字（PNG 图片）
        'BM': ('sample_coord_x', 'input'),          # 打卡坐标x（回退 extras 打卡）
        'BN': ('sample_coord_y', 'input'),          # 打卡坐标y
        'BO': ('track', 'extra'),                   # 轨迹 GPX 文件名
        'BP': ('photos', 'extra'),                  # 小班照片文件名
    },
    "table3": {  # 表3-退化林修复验收因子表（新布局：M造林树种 O补植面积 AA修复措施 AB修复方式 AC辅助措施 AU小班设计株树）
        'A': ('city', 'prefilled'),
        'B': ('subcompartment', 'prefilled'),       # 调查小班号
        'C': ('county', 'prefilled'),
        'D': ('township', 'prefilled'),
        'E': ('village', 'prefilled'),
        'F': ('forest_compartment', 'prefilled'),
        'G': ('subcompartment_orig', 'prefilled'),  # 小班（GDB 原值，可含汉字；无则回退调查小班号）
        'H': ('check_type', 'prefilled'),
        'I': ('project_name', 'prefilled'),
        'J': ('plan_year', 'prefilled'),
        'K': ('work_year', 'prefilled'),
        'L': ('ownership', 'prefilled'),
        'M': ('tree_species', 'prefilled'),         # 造林树种（GDB 绿色）
        'N': ('reported_area', 'prefilled'),
        'O': ('replant_area', 'prefilled'),         # 补植面积（GDB 绿色）
        'P': ('survival_pass', 'input'),
        'Q': ('survival_replant', 'input'),
        'R': ('survival_fail', 'input'),
        'S': ('verified_total', 'input'),
        'T': ('verified_pass', 'input'),
        'U': ('verified_replant', 'input'),
        'V': ('verified_fail', 'input'),
        'W': ('verified_loss', 'input'),
        'X': ('area_short_reason', 'input'),
        'Y': ('unqualified_reason', 'input'),
        'Z': ('loss_reason', 'input'),
        'AA': ('repair_measure', 'prefilled'),      # 修复措施（GDB 绿色）
        'AB': ('repair_method', 'prefilled'),       # 修复方式（GDB 绿色）
        'AC': ('auxiliary_measure', 'prefilled'),   # 辅助措施（GDB 绿色）
        'AD': ('mgmt_design', 'input'),
        'AE': ('mgmt_meeting', 'input'),
        'AF': ('mgmt_speech', 'input'),
        'AG': ('mgmt_survey', 'input'),
        'AH': ('mgmt_supervision', 'input'),
        'AI': ('mgmt_photo', 'input'),
        'AJ': ('construction_area', 'input'),
        'AK': ('construction_rate', 'input'),
        'AL': ('archive_area', 'input'),
        'AM': ('archive_rate', 'input'),
        'AN': ('protect_area', 'input'),
        'AO': ('protect_rate', 'input'),
        'AP': ('tend_area', 'input'),
        'AQ': ('tend_rate', 'input'),
        'AR': ('planted_total', 'sample_stat'),     # 小班查数株数 = 调查总株数（B34 口径）
        'AS': ('qualified_count', 'sample_stat'),   # 合格株树
        'AT': ('qualified_rate', 'sample_stat'),    # 合格率
        'AU': ('design_count', 'prefilled'),        # 小班设计株树（GDB 绿色）
        'AV': ('inspector_sign', 'sign'),           # 验收人员手写签字（PNG 图片）
        'AW': ('inspect_time', 'input'),
        'AX': ('remark', 'input'),
        'AY': ('co_inspector_sign', 'sign'),        # 配合验收人员手写签字（PNG 图片）
        'AZ': ('sample_coord_x', 'input'),          # 打卡坐标x（回退 extras 打卡）
        'BA': ('sample_coord_y', 'input'),          # 打卡坐标y
        'BB': ('track', 'extra'),                   # 轨迹 GPX 文件名
        'BC': ('photos', 'extra'),                  # 小班照片文件名
    },
}

# 分类 → 基本信息模板 sheet 名（tpl-base.xlsx 内）
_TPL_SHEET_NAMES = {
    "人工造林": "2023年度人工造林",
    "封山育林": "2023年度封山育林",
    "退化林修复": "2023年度退化林",
}

# 基本信息模板数据起始行（1-2 标题，3-4 表头）
_BASE_DATA_START = 5

# 样地模板块结构（tpl-samples.xlsx，块高 39 行）
_SAMPLE_BLOCK_ROWS = 39
_SAMPLE_DATA_START = 4    # 块内数据起始行（样地号 1）
_SAMPLE_DATA_SLOTS = 23   # 模板预留样地槽位数（行4-26）

# 公式内单元格引用（$A$1 / A1 形式），块复制时行号偏移用
_CELL_RE = re.compile(r'(\$?[A-Z]{1,2}\$?)(\d+)')

# 率类列（百分比格式 0.00%）：合格率 + 成活率等级3列 + 施工/建档/管护/抚育率4列。
# 模板仅 R5 示例行带格式，导出写行 6+ 时单元格为新建（General），
# 写值/写分派公式时须补设 number_format，否则比率 0.9524 显示为裸数字。
_RATE_COL_FMT = '0.00%'
_RATE_COLS = {
    "table1": {"O", "P", "Q", "AP", "AG", "AI", "AK", "AM"},
    "table2": {"U", "V", "W", "BG", "AX", "AZ", "BB", "BD"},
    "table3": {"P", "Q", "R", "AT", "AK", "AM", "AO", "AQ"},
}


# ════════════════════════════════════════════
# 通用工具
# ════════════════════════════════════════════

def _num(v):
    """安全转 int（排序用）：空/None/无效 → 0。"""
    try:
        return int(float(v or 0))
    except (TypeError, ValueError):
        return 0


def _sc_sort_key(r):
    """小班行排序键：(林班, 调查小班号) 数字序。"""
    return (_num(r.get("forest_compartment")), _num(r.get("subcompartment")))


def _col_letter(n):
    """数字转 Excel 列字母：1→A, 27→AA"""
    result = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def _col_number(letter):
    """Excel 列字母转数字：A→1, AA→27"""
    n = 0
    for c in letter.upper():
        n = n * 26 + (ord(c) - 64)
    return n


def _norm_enum(val, f):
    """enum 字段归一：空值取默认值；旧 checkbox 布尔数据映射为选项。"""
    if f.get("type") != "enum" or not f.get("options"):
        return val
    if val in ("", None):
        return f.get("default", "")
    if val is True or val == "true":
        return f["options"][0]
    if val is False or val == "false" or val == 0 or val == "0":
        return f["options"][-1] if len(f["options"]) == 2 else f.get("default", "")
    return val


def _fmt_num(val):
    """数值化导出：'85.50' → 85.5，'85' → 85（去尾零，最多保留2位小数）。"""
    try:
        num = float(val)
    except (ValueError, TypeError):
        return val
    if num == int(num):
        return int(num)
    return round(num, 2)


def _sample_stats(data):
    """样地聚合统计（苗木合格率组），口径与样地页汇总/模板 B34 一致。

    小班查数株数 = 调查总株数 = round(Σ种植 ÷ 个数 ÷ 150 × 单个网格面积 × 种植网格数量)
                  个数 = 手写 sm_total_count（>0 生效）回退实际样地数；
                  个数或Σ种植为 0 → 空（同模板 IF(OR(B27=0,B29=0),"",…) 守卫），
                  网格面积/数量未填 → 结果为 0（同模板 B32/B33 空参与乘法）
    合格率      = Σ成活÷Σ种植（比率 0-1 如 0.9524，round 4 位；模板合格率列
                  已设 0.00% 百分比格式，Excel 显示 95.24%——不 ×100）
    合格株树    = round(查数株数×合格率)
    """
    data = data if isinstance(data, dict) else {}
    samples = data.get("samples")

    def _f(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0

    planted = 0.0
    alive = 0.0
    real_n = 0
    for s in samples if isinstance(samples, list) else []:
        if not isinstance(s, dict):
            continue
        real_n += 1
        planted += _f(s.get("planted"))
        alive += _f(s.get("alive"))
    if not planted:
        return {"planted_total": None, "qualified_count": None, "qualified_rate": None}
    # 合格率 = 比率 0-1（不 ×100）：模板合格率列/成活率等级列均为 0.00%
    # 百分比格式、分派公式阈值 0.9/0.4 按比率比较（CONSTRAINTS C8/C9/D3）
    rate = round(alive / planted, 4)
    # 个数：手写 sm_total_count 优先（与 B27/前端 smSummaryComputed 同口径）
    n = None
    hand_n = data.get("sm_total_count")
    if hand_n not in ("", None):
        try:
            hf = float(hand_n)
            if hf > 0:
                n = int(round(hf))
        except (ValueError, TypeError):
            n = None
    if n is None:
        n = real_n
    total = round(planted / n / 150 * _f(data.get("sm_grid_area"))
                  * _f(data.get("sm_grid_count"))) if n > 0 else None
    return {
        "planted_total": total,
        "qualified_rate": rate,
        "qualified_count": round(total * rate) if total is not None else None,
    }


def _shift_formula(formula, offset):
    """公式内单元格引用行号 +offset（块复制用）。

    如 '=SUM(B5:B27)' offset=39 → '=SUM(B44:B66)'；
    函数名（ROUND/SUM 后跟括号）与纯数字常量（150、,0）不受影响。
    """
    return _CELL_RE.sub(lambda m: f"{m.group(1)}{int(m.group(2)) + offset}", formula)


# ════════════════════════════════════════════
# 基本信息导出（tpl-base）
# ════════════════════════════════════════════

def _resolve_cell(key, source, prefilled, input_data, extras, stats, sc_row, field_types):
    """按 (key, source) 解析一个单元格的值。"""
    if source == 'prefilled':
        if key == 'subcompartment_orig':
            # 小班：GDB 原值（可含汉字如「红9」，_fmt_num 转换失败原样保留）；
            # 无「小班」字段时回退调查小班号；数值串 "1.0"→1
            val = prefilled.get('subcompartment_orig') or prefilled.get('subcompartment', '')
            return _fmt_num(val) if val not in ('', None) else ''
        if key == 'forest_compartment':
            # 林班 0 = 无林班（GDB 空），导出留空与模板示例一致
            v = prefilled.get('forest_compartment', '')
            return '' if not _num(v) else v
        val = prefilled.get(key, '')
        # 数值化（面积/株数/年度等）：GDB 属性以字符串入库（"2023.0"/"11000.0"），
        # 整数值浮点去 ".0"（小班/年度/株数不可能有小数）；含汉字的原始小班号
        # 等非数值串 float() 失败原样保留
        if val not in ('', None):
            return _fmt_num(val)
        return val
    if source == 'input':
        # 打卡坐标：录入值优先，回退 extras 打卡记录。
        # 坐标写全精度浮点（GPS 约 6-7 位小数），不走 _fmt_num 的 2 位舍入
        if key in ('sample_coord_x', 'sample_coord_y'):
            v = input_data.get(key)
            if v in ('', None):
                ek = 'checkin_lng' if key == 'sample_coord_x' else 'checkin_lat'
                v = (extras or {}).get(ek, '')
            if v in ('', None):
                return ''
            try:
                return float(v)
            except (ValueError, TypeError):
                return v
        val = input_data.get(key, '')
        f = field_types.get(key)
        if f:
            t = f.get('type')
            if t == 'enum':
                # 空值取默认：管理情况 5 项默认「有」——前端不保存未触碰的
                # 默认值，库里缺 key，导出时按 schema 默认补齐
                return _norm_enum(val, f)
            if val not in ('', None):
                if t == 'percent':
                    # 率类存比率 0-1（不 ×100）：保留 4 位小数（0.9524），
                    # 配合模板 0.00% 百分比格式显示 95.24%——不走 _fmt_num
                    # 的 2 位舍入（0.9524 → 0.95 会丢精度）
                    try:
                        return round(float(val), 4)
                    except (ValueError, TypeError):
                        return val
                if t == 'number':
                    return _fmt_num(val)
                if t == 'photo':
                    if isinstance(val, list):
                        return ';'.join(str(x) for x in val if x)
        return val
    if source == 'sign':
        # 手写签字：data_json 里的 PNG data URL（由前端签字 canvas 保存），
        # 值本身不写单元格——由填充循环调 _insert_sign_image 以图片插入
        return input_data.get(key, '')
    if source == 'sample_stat':
        return stats.get(key)
    if source == 'extra':
        if key == 'track':
            return _track_gpx_filename(sc_row) if (extras or {}).get('track') else ''
        if key == 'photos':
            names = [p.get('name', '') for p in ((extras or {}).get('photos') or [])
                     if isinstance(p, dict) and p.get('name')]
            return ';'.join(names)
    return ''


def _insert_sign_image(ws, row, col, data_url):
    """把手写签字 PNG（data URL）插入单元格：裁白边 → 等比缩放至固定范围 → 锚定格左上角。

    签字 canvas 大部分是空白，先裁掉白边只留笔迹包围盒；再等比缩放到
    固定显示范围（宽 ≤200px、高 ≤64px，只缩不放）——不按单元格尺寸
    计算，允许图片浮在格上（Excel/WPS 中点击图片可放大查看原图）。
    无笔迹/解析失败静默跳过（不影响导出）。
    """
    try:
        from PIL import Image as PILImage, ImageOps
        from openpyxl.drawing.image import Image as XLImage
    except ImportError:
        return  # 环境缺 Pillow：跳过签字图片（其余导出不受影响）
    try:
        b64 = data_url.split(',', 1)[1]
        raw = base64.b64decode(b64)
        im = PILImage.open(io.BytesIO(raw))
        # 裁白边：反色后 getbbox 得到笔迹包围盒（canvas 白底黑迹）
        bbox = ImageOps.invert(im.convert('L')).getbbox()
        if bbox:
            pad = 4
            im = im.crop((max(0, bbox[0] - pad), max(0, bbox[1] - pad),
                          min(im.width, bbox[2] + pad), min(im.height, bbox[3] + pad)))
        if im.width < 8 or im.height < 8:
            return  # 空白签字
        # 固定显示范围等比缩放（只缩不放）
        max_w, max_h = 200, 64
        scale = min(max_w / im.width, max_h / im.height, 1.0)
        buf = io.BytesIO()
        im.save(buf, format='PNG')
        buf.seek(0)
        img = XLImage(buf)
        img.width = max(1, int(im.width * scale))
        img.height = max(1, int(im.height * scale))
        img.anchor = ws.cell(row=row, column=col).coordinate  # 锚定单元格左上角
        ws.add_image(img)
    except Exception:
        return  # 单条签字损坏不阻断整个导出


def _clear_base_data_region(ws):
    """清空基本信息模板数据区（行5起），去除官方模板预填的示例数据。

    保留 1-4 行表头与行样式/合并单元格，仅清单元格值。
    官方 tpl-base.xlsx 预填了华宁/澄江等示例行（最多 788 行），
    不清空会导致导出文件残留无关示例数据。
    """
    for r in range(_BASE_DATA_START, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).value = None


def export_base(pid, output_path=None, category=None):
    """导出项目基本信息 xlsx（tpl-base 模板，一项目一文件，3 分类 sheet）。

    每个分类 sheet 填入该项目该分类下全部小班（含未录入的小班，
    白色列留空），行序按（林班, 调查小班号）数字序。

    Args:
        pid: 项目 ID
        output_path: 输出路径，None 则返回 BytesIO
        category: 可选，仅导出该分类（用户端 topbar「导出」按钮按当前分类导出）。
                  指定时输出文件只保留该分类 sheet，其余分类 sheet 移除。

    Returns:
        (output_path, stats) 元组
    """
    check_templates()
    project = storage.get_project(pid)
    if not project:
        raise ValueError(f"项目 {pid} 不存在")
    if category and category not in _TPL_SHEET_NAMES:
        raise ValueError(f"未知分类「{category}」")
    wb = openpyxl.load_workbook(_BASE_TEMPLATE)
    stats = {"project": project["name"], "sheets": {}}

    for cat, sheet_name in _TPL_SHEET_NAMES.items():
        # 分类过滤：只保留当前分类 sheet，其余移除
        if category and cat != category:
            wb.remove(wb[sheet_name])
            continue
        table_id = GDB.GDB_CATEGORY_TO_TABLE.get(cat)
        table_def = S.get_table(table_id)
        if not table_def:
            continue
        ws = wb[sheet_name]
        # 快照模板示例行（R5）的公式：清空数据区会连公式一起清掉。
        # 模板 2026-08-21 起自带成活率/合格率分派公式（O/P/Q/S… 列），
        # 写行时按行偏移代入每个数据行；exporter 有值的列仍以录入值优先。
        tpl_formulas = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=_BASE_DATA_START, column=c).value
            if isinstance(v, str) and v.startswith("="):
                tpl_formulas[c] = v
        _clear_base_data_region(ws)  # 去官方模板预填示例
        sc_rows = storage.list_project_subcompartment_rows(pid, category=cat)
        sc_rows.sort(key=_sc_sort_key)
        survey_map = {rec["subcompartment_id"]: rec.get("data", {}) or {}
                      for rec in storage.get_survey_rows(pid, table_id)}
        # schema 字段类型索引（enum/number/percent/photo 转换用）
        field_types = {f["key"]: f for f in table_def.get("input_columns", [])}
        need_extra = any(src == 'extra' or (key or '').startswith('sample_coord')
                         for _, (key, src) in _TPL_COL_MAPS[table_id].items())

        rate_cols = _RATE_COLS.get(table_id, set())
        # store:false 派生列（成活率等级/面积分派）：不写录入值，
        # 一律由模板公式 =IF(AP>=0.9,…) 现算（历史残留值不覆盖公式）
        no_store_keys = {f["key"] for f in table_def.get("input_columns", [])
                         if f.get("type") == "computed" and f.get("store") is False}
        for i, sc_row in enumerate(sc_rows):
            row_num = _BASE_DATA_START + i
            prefilled = S.map_subcompartment_to_prefilled(sc_row.get("data", {}))
            input_data = survey_map.get(sc_row["id"], {}) or {}
            stats_v = _sample_stats(input_data)
            extras = storage.get_extras(sc_row["id"]) if need_extra else None
            for col_letter, (key, source) in _TPL_COL_MAPS[table_id].items():
                if key in no_store_keys:
                    continue
                val = _resolve_cell(key, source, prefilled, input_data,
                                    extras, stats_v, sc_row, field_types)
                if source == 'sign':
                    # 手写签字：data URL → 裁白边缩放后以图片插入（不写文本）
                    if isinstance(val, str) and val.startswith('data:image'):
                        _insert_sign_image(ws, row_num, _col_number(col_letter), val)
                    continue
                if val not in ('', None):
                    cell = ws.cell(row=row_num, column=_col_number(col_letter), value=val)
                    if col_letter in rate_cols:
                        cell.number_format = _RATE_COL_FMT
            # 模板公式代入下一行：该行该列没写值的，填行偏移后的模板公式
            # （如 R6 的 O 列 = IF(AP6>=0.9,AP6,"")，由 _shift_formula 平移行号）
            for c, f in tpl_formulas.items():
                cell = ws.cell(row=row_num, column=c)
                if cell.value in (None, ""):
                    cell.value = _shift_formula(f, row_num - _BASE_DATA_START)
                    if _col_letter(c) in rate_cols:
                        cell.number_format = _RATE_COL_FMT
        stats["sheets"][cat] = len(sc_rows)

    if output_path is None:
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output, stats
    wb.save(str(output_path))
    return str(output_path), stats


# ════════════════════════════════════════════
# 样地导出（tpl-samples，块结构）
# ════════════════════════════════════════════

def _copy_sample_block(ws, offset):
    """把模板块1（行1-38）复制到 offset+1 起的位置。

    复制值（公式行号偏移）、单元格样式、行高、块内合并单元格。
    """
    dst_start = offset + 1
    # 值 + 样式 + 行高
    for r in range(_SAMPLE_BLOCK_ROWS):
        src_r, dst_r = 1 + r, dst_start + r
        h = ws.row_dimensions[src_r].height
        if h is not None:
            ws.row_dimensions[dst_r].height = h
        for c in range(1, ws.max_column + 1):
            src = ws.cell(row=src_r, column=c)
            dst = ws.cell(row=dst_r, column=c)
            v = src.value
            if isinstance(v, str) and v.startswith('='):
                v = _shift_formula(v, offset)
            dst.value = v
            if src.has_style:
                dst._style = _style_copy(src._style)
    # 合并单元格（仅块1内的 range）
    for m in list(ws.merged_cells.ranges):
        if 1 <= m.min_row and m.max_row <= _SAMPLE_BLOCK_ROWS:
            ws.merge_cells(start_row=m.min_row + offset, start_column=m.min_col,
                           end_row=m.max_row + offset, end_column=m.max_col)


def _fill_sample_block(ws, block_idx, cat, project_name, sc_row, data):
    """填充一个小班的样地块（行 block_idx*39+1 起）。

    data 为该小班表记录的 data_json：samples 数组 + sm_* 汇总手写项。
    """
    base = block_idx * _SAMPLE_BLOCK_ROWS
    if block_idx > 0:
        _copy_sample_block(ws, base)
    prefilled = S.map_subcompartment_to_prefilled(sc_row.get("data", {}))
    samples = data.get("samples", []) if isinstance(data, dict) else []
    samples = samples if isinstance(samples, list) else []

    # R1 标题（项目名称+类型+调查小班号）；R2 年度县乡；R3 列头（不动）
    # gdb_data：GDB 导入的小班属性（区别于参数 data＝调查记录 data_json）
    gdb_data = sc_row.get("data", {}) or {}

    def _int_like(v):
        """2023.0 → 2023、5.0 → 5（GDB 浮点字段按整数显示）。"""
        try:
            f = float(v)
            return str(int(f)) if f == int(f) else str(f)
        except (ValueError, TypeError):
            return str(v or "").strip()

    # 小班原值：新导入读「小班」（GDB 原值，可含汉字）；旧数据兼容读「小班原始」；
    # 均无回退调查小班号（相同则即调查小班号本身）
    orig_no = _int_like(gdb_data.get("小班原始") or gdb_data.get("小班")
                        or sc_row.get("subcompartment"))
    survey_no = _int_like(sc_row.get("subcompartment"))
    # 标题含调查小班号（模板占位「样地调查表（项目名称+类型）调查小班号」）；
    # 原始号≠调查号时括注保留小班原始号信息（D9 口径）
    title = f"{project_name}样地调查表（{cat}）调查小班号{survey_no}"
    if orig_no != survey_no:
        title += f"（小班{orig_no}）"
    ws.cell(row=base + 1, column=1, value=title)
    # R2 年度 县区（模板 A2:E2 合并；F2 坐标 / H2 照片 为分组表头不动）
    year = _int_like(prefilled.get("plan_year") or prefilled.get("work_year"))
    parts = []
    if year:
        parts.append(f"{year}年度")
    for k in ("county", "township"):
        v = prefilled.get(k, "")
        if v not in ("", None):
            parts.append(str(v))
    if parts:
        ws.cell(row=base + 2, column=1, value="  ".join(parts))

    # 样地数据槽（行4-26，最多23个）：先清残留的 E 列公式（块1模板自带 R4-R7），
    # 避免无数据槽位在 Excel 中显示 0
    # （注意：ws.cell(..., value=None) 不会清值，必须属性赋值）
    for j in range(_SAMPLE_DATA_SLOTS):
        ws.cell(row=base + _SAMPLE_DATA_START + j, column=5).value = None
    for j, s in enumerate(samples[:_SAMPLE_DATA_SLOTS] if isinstance(samples, list) else []):
        if not isinstance(s, dict):
            continue
        r = base + _SAMPLE_DATA_START + j
        ws.cell(row=r, column=1, value=s.get("no") or (j + 1))
        for col, key in ((2, 'area'), (3, 'planted'), (4, 'alive')):
            v = s.get(key)
            if v not in ('', None):
                ws.cell(row=r, column=col, value=_fmt_num(v))
        # 坐标（x/y）：写原始浮点值不做四舍五入——GPS 精度约 6-7 位小数，
        # _fmt_num 的 2 位舍入会把 102.663521 截成 102.66 导致点位偏移
        for col, key in ((6, 'x'), (7, 'y')):
            v = s.get(key)
            if v not in ('', None):
                try:
                    ws.cell(row=r, column=col, value=float(v))
                except (ValueError, TypeError):
                    ws.cell(row=r, column=col, value=v)
        # 死亡株数 = 种植-成活（模板公式口径，统一补写保证每行都有）
        ws.cell(row=r, column=5, value=f"=C{r}-D{r}")
        # 样地照片（仅文件名，不上传）
        photos = s.get("photos") or []
        if photos:
            names = [str(p) for p in (photos if isinstance(photos, list) else [photos]) if p]
            if names:
                ws.cell(row=r, column=8, value=";".join(names))
        # 样地备注（I 列，文本；样地页卡片成活株数后的「备注」输入框）
        rv = s.get("remark")
        if rv not in ('', None):
            ws.cell(row=r, column=9, value=str(rv))

    # 总样地个数（B27）：优先手写值（样地页汇总表单 sm_total_count，前端默认
    # 自动填充实际样地数、手改后以用户值为准；与前端 smSummaryComputed 同口径：
    # 手写值 >0 才生效），无有效手写值回退实际样地数；
    # B28-B30 SUM / B31 成活率 / B34 调查总株数 公式由块复制保留（模板已加除0守卫，
    # B34 公式引用 B27，手写个数同样参与计算）
    n = None
    hand_n = data.get('sm_total_count') if isinstance(data, dict) else None
    if hand_n not in ('', None):
        try:
            nf = float(hand_n)
            if nf > 0:
                n = int(round(nf))
        except (ValueError, TypeError):
            n = None
    if n is None:
        n = sum(1 for s in samples if isinstance(s, dict))
    ws.cell(row=base + 27, column=2, value=n)

    # 手写录入项（样地页汇总表单，data_json.sm_* 键）：
    # B32 单个网格面积 / B33 种植网格数量（数值，参与 B34 公式计算）
    for r, k in ((32, 'sm_grid_area'), (33, 'sm_grid_count')):
        v = data.get(k) if isinstance(data, dict) else None
        if v not in ('', None):
            try:
                ws.cell(row=base + r, column=2, value=float(v))
            except (ValueError, TypeError):
                ws.cell(row=base + r, column=2, value=v)
    # B35 撑杆 / B36 覆膜 / B37 验收人 / B38 验收日期 / B39 备注（文本）
    for r, k in ((35, 'sm_pole'), (36, 'sm_film'), (37, 'sm_inspector'),
                 (38, 'sm_inspect_date'), (39, 'sm_remark')):
        v = data.get(k) if isinstance(data, dict) else None
        if v not in ('', None):
            ws.cell(row=base + r, column=2, value=str(v))


def _clear_sample_template_block(ws):
    """清空样地模板块1的预填示例，保留标题/表头/汇总公式。

    样地模板 tpl-samples.xlsx 块1预填了示例样地（R4-R7 的 150/22/21 等）、
    总样地个数(R27 B=5)、单个网格面积(R32 B=5000)、种植网格数量(R33 B=4)。
    不清空会残留在导出文件里。块1清空后，copy_worksheet 得到的每个块都是干净脚手架。
    汇总公式（R28-R31：=SUM/=B30/B29 等、R34=ROUND）保留不动。
    """
    # 先拆掉样地数据区（行4-26）内的合并单元格：数据行 A-H 每格独立，
    # 模板被手动编辑后可能残留合并（如 B27:H27），不拆会导致写值报
    # 'MergedCell' object attribute 'value' is read-only
    for m in list(ws.merged_cells.ranges):
        if m.min_row >= _SAMPLE_DATA_START and m.max_row < _SAMPLE_DATA_START + _SAMPLE_DATA_SLOTS:
            ws.unmerge_cells(start_row=m.min_row, start_column=m.min_col,
                             end_row=m.max_row, end_column=m.max_col)
    # 样地槽（行4-26）整行 A-I 清空（I 列为样地备注，2026-08-21 起录入导出）
    for r in range(_SAMPLE_DATA_START, _SAMPLE_DATA_START + _SAMPLE_DATA_SLOTS):
        for c in range(1, 10):  # A-I
            ws.cell(row=r, column=c).value = None
    # 手写项残留清理：总样地个数(R27 B)/单个网格面积(R32 B)/种植网格数量(R33 B)
    for r in (27, 32, 33):
        ws.cell(row=r, column=2).value = None


def _sheet_safe(name):
    """sheet 名合法化：去掉 Excel 禁止字符（: \\ / ? * [ ]）并截断到 31 字符。"""
    cleaned = re.sub(r"[:\\/?*\[\]]", "", str(name)).strip()
    return (cleaned or "sheet")[:31]


def export_samples(pid, output_path=None, subcompartment_id=None, category=None):
    """导出项目样地 xlsx（tpl-samples 模板，每分类一个 sheet，每小班一个 39 行块）。

    Args:
        pid: 项目 ID
        output_path: 输出路径，None 则返回 BytesIO
        subcompartment_id: 可选，仅导出该小班（用户端样地页「导出」按钮）。
                           单小班模式下 sheet 名为「分类-调查小班号」。
        category: 可选，仅导出该分类（admin 项目管理「分类下载」，每分类一个文件）。

    Returns:
        (output_path, stats) 元组
    """
    check_templates()
    project = storage.get_project(pid)
    if not project:
        raise ValueError(f"项目 {pid} 不存在")
    wb = openpyxl.load_workbook(_SAMPLE_TEMPLATE)
    tpl_ws = wb[wb.sheetnames[0]]
    _clear_sample_template_block(tpl_ws)  # 去官方模板预填示例，保证每块脚手架干净
    stats = {"project": project["name"], "sheets": {}}

    if subcompartment_id:
        # 单小班模式：仅当前小班、当前分类，sheet 名加调查小班号
        sc_row = storage.get_subcompartment_row(subcompartment_id)
        # 项目归属校验：subcompartment_rows 无 project_id 列，经 gdb_files 关联
        gdb = storage.get_gdb_file(sc_row.get("gdb_id")) if sc_row else None
        if not sc_row or not gdb or gdb.get("project_id") != pid:
            raise ValueError("小班不存在")
        cat = sc_row.get("category") or ""
        table_id = GDB.GDB_CATEGORY_TO_TABLE.get(cat)
        if not table_id:
            raise ValueError(f"小班分类「{cat}」无对应调查表")
        sheet_name = _sheet_safe(
            f"{cat}-{sc_row.get('subcompartment') or sc_row.get('subcompartment_label') or ''}"
        )
        ws = wb.copy_worksheet(tpl_ws)
        ws.title = sheet_name
        survey_map = {rec["subcompartment_id"]: rec.get("data", {}) or {}
                      for rec in storage.get_survey_rows(pid, table_id)}
        data = survey_map.get(sc_row["id"], {}) or {}
        _fill_sample_block(ws, 0, cat, project["name"], sc_row, data)
        stats["sheets"][sheet_name] = 1
    else:
        for cat in ([category] if category else list(_TPL_SHEET_NAMES)):
            table_id = GDB.GDB_CATEGORY_TO_TABLE.get(cat)
            if not table_id:
                continue
            sc_rows = storage.list_project_subcompartment_rows(pid, category=cat)
            sc_rows.sort(key=_sc_sort_key)
            if not sc_rows:
                continue
            ws = wb.copy_worksheet(tpl_ws)
            ws.title = cat
            survey_map = {rec["subcompartment_id"]: rec.get("data", {}) or {}
                          for rec in storage.get_survey_rows(pid, table_id)}
            for i, sc_row in enumerate(sc_rows):
                data = survey_map.get(sc_row["id"], {}) or {}
                _fill_sample_block(ws, i, cat, project["name"], sc_row, data)
            stats["sheets"][cat] = len(sc_rows)

    wb.remove(tpl_ws)
    if not stats["sheets"]:
        raise ValueError("该项目暂无小班数据")

    if output_path is None:
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output, stats
    wb.save(str(output_path))
    return str(output_path), stats


def export_samples_zip(pid, category=None):
    """样地按小班拆分导出并打包 zip（后台「分类下载」）。

    每小班一个 xlsx（单小班模式：该分类一个 sheet、该小班一个块），
    文件名 {林班-小班|小班}号调查小班-{分类}-{年度}.xlsx；
    年度取小班计划年度（GDB）→ 项目名「(2023 年度)」→ 当前年；
    同名冲突（跨村重复林班小班号）追加 _2/_3 序号。

    Returns:
        (BytesIO(zip), {"files": n})
    """
    import zipfile
    sc_rows = storage.list_project_subcompartment_rows(pid, category=category)
    sc_rows.sort(key=_sc_sort_key)
    if not sc_rows:
        raise ValueError("该分类暂无小班数据")
    buf = io.BytesIO()
    used = {}
    n = 0
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for sc_row in sc_rows:
            try:
                out, _stats = export_samples(pid, subcompartment_id=sc_row["id"])
            except ValueError:
                continue  # 分类无对应调查表等异常小班跳过
            base = _sc_file_base(sc_row)
            k = used.get(base, 0) + 1
            used[base] = k
            zf.writestr(f"{base}.xlsx" if k == 1 else f"{base}_{k}.xlsx",
                        out.getvalue())
            n += 1
    if n == 0:
        raise ValueError("该分类暂无样地数据")
    buf.seek(0)
    return buf, {"files": n}


# ════════════════════════════════════════════
# 轨迹 GPX 导出（ArcGIS Pro / ArcMap / QGIS 直接识别）
# ════════════════════════════════════════════

def _parse_year(v):
    """年度值清洗："2023.0"/2023/"2023" → "2023"；无法解析返回 None。"""
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    try:
        n = float(s)
    except ValueError:
        return None
    return str(int(round(n))) if n > 0 else None


def _sc_year(sc_row):
    """小班年度：GDB 计划年度 → 所属项目名「(2023 年度)」正则 → 当前年。"""
    data = sc_row.get("data")
    y = _parse_year(data.get("计划年度")) if isinstance(data, dict) else None
    if y:
        return y
    m = re.search(r"(\d{4})\s*年度", str(sc_row.get("project_name") or ""))
    if m:
        return m.group(1)
    return str(_date.today().year)


def _sc_file_stem(sc_row):
    """小班文件名主干：林班>0 时「林班-小班」（防跨林班同号冲突），否则小班号。"""
    try:
        fc = int(sc_row.get("forest_compartment") or 0)
    except (TypeError, ValueError):
        fc = 0
    sc = sc_row.get("subcompartment")
    if sc in (None, ""):
        sc = sc_row.get("subcompartment_label")
    return f"{fc}-{sc}" if fc > 0 else f"{sc}"


def _sc_file_base(sc_row):
    """导出文件名主干（去扩展名）：{林班-小班|小班}号调查小班-{分类}-{年度}。"""
    return "-".join(str(p).strip() for p in (
        f"{_sc_file_stem(sc_row)}号调查小班",
        sc_row.get("category") or "",
        _sc_year(sc_row),
    ) if p and str(p).strip())


def _track_gpx_filename(sc_row):
    """轨迹 GPX 文件名：{林班-小班|小班}号调查小班-{分类}-{年度}.gpx。

    2026-08-21 改版（与后台分类下载的样地文件命名统一）：
    旧名「分类_乡镇_村_小班.gpx」→ 新名如「1号调查小班-人工造林-2023.gpx」。
    Excel 轨迹列与 GPX zip 内文件名同源（本函数），改版后两处一致。
    """
    return _sc_file_base(sc_row) + ".gpx"


def _track_gpx(track, name):
    """轨迹点列表 → GPX 1.1 文本。track = [{lng, lat, t}, ...]（WGS84）。"""
    pts = []
    for p in track:
        try:
            lng = float(p.get("lng"))
            lat = float(p.get("lat"))
        except (TypeError, ValueError):
            continue
        t = str(p.get("t") or "").strip()
        time_el = f"<time>{t}</time>" if t else ""
        pts.append(f'      <trkpt lat="{lat}" lon="{lng}">{time_el}</trkpt>')
    if not pts:
        return ""
    return (
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        '<gpx version="1.1" creator="hqz-survey" '
        'xmlns="http://www.topografix.com/GPX/1/1">\n'
        f'  <trk>\n    <name>{name}</name>\n    <trkseg>\n'
        + "\n".join(pts)
        + '\n    </trkseg>\n  </trk>\n</gpx>\n'
    )


def _track_coords(track):
    """轨迹点 → [(lng, lat), ...]（跳过无效点，WGS84）。"""
    coords = []
    for p in track:
        try:
            coords.append((float(p.get("lng")), float(p.get("lat"))))
        except (TypeError, ValueError):
            continue
    return coords


def _track_times(track):
    """轨迹点时间戳列表（非空 t，保持顺序）。"""
    return [str(p.get("t") or "").strip() for p in track
            if str(p.get("t") or "").strip()]


def _proj_year(proj):
    """项目年度：项目名「(2023 年度)」正则 → 当前年（与 admin _dl_year 同口径）。"""
    m = re.search(r"(\d{4})\s*年度", (proj or {}).get("name") or "")
    return m.group(1) if m else str(_date.today().year)


def _dl_stem(proj, category):
    """单文件轨迹（KML/SHP）文件主干：{分类}-{年度} 或 项目名（清路径非法字符）。"""
    stem = f"{category}-{_proj_year(proj)}" if category else \
        (proj or {}).get("name") or "轨迹"
    return re.sub(r'[\\/:*?"<>|]+', "_", str(stem)).strip() or "轨迹"


def _track_desc(sc_row, track):
    """KML Placemark 描述：人类可读的小班摘要。"""
    data = sc_row.get("data") if isinstance(sc_row.get("data"), dict) else {}
    ts = _track_times(track)
    bits = [
        f"分类: {sc_row.get('category') or '—'}",
        f"小班: {sc_row.get('subcompartment_label') or _sc_file_stem(sc_row)}",
        f"乡镇: {data.get('乡镇') or '—'}",
        f"村: {data.get('村') or '—'}",
        f"点数: {len(track)}",
    ]
    if ts:
        bits.append(f"起止: {ts[0]} ~ {ts[-1]}")
    return " | ".join(bits)


def _track_kml(doc_name, entries):
    """entries = [(name, desc, coords)] → KML 2.2 文本。

    全部小班各一个 Placemark（LineString）；ArcGIS 用 KML To Layer
    转要素，Google Earth / 奥维可直接打开。
    """
    from xml.sax.saxutils import escape as _xml_escape
    parts = []
    for name, desc, coords in entries:
        co = " ".join(f"{lng},{lat}" for lng, lat in coords)
        d = (f"\n    <description>{_xml_escape(desc)}</description>"
             if desc else "")
        parts.append(
            f"  <Placemark>\n    <name>{_xml_escape(name)}</name>{d}\n"
            f"    <LineString><tessellate>1</tessellate>"
            f"<coordinates>{co}</coordinates></LineString>\n  </Placemark>")
    if not parts:
        return ""
    return (
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        '<kml xmlns="http://www.opengis.net/kml/2.2">\n'
        f'  <Document>\n    <name>{_xml_escape(doc_name)}</name>\n'
        + "\n".join(parts) + "\n  </Document>\n</kml>\n"
    )


def _track_shp_props(sc_row, track):
    """小班 → shapefile 属性行。字段名限 ASCII ≤10 字符（DBF 规范，ArcGIS 10 兼容）。"""
    data = sc_row.get("data") if isinstance(sc_row.get("data"), dict) else {}
    ts = _track_times(track)
    try:
        fc = int(sc_row.get("forest_compartment") or 0)
    except (TypeError, ValueError):
        fc = 0
    try:
        sc = int(sc_row.get("subcompartment") or 0)
    except (TypeError, ValueError):
        sc = None
    return {
        "name": sc_row.get("subcompartment_label") or _sc_file_stem(sc_row),
        "category": sc_row.get("category") or "",
        "year": _sc_year(sc_row),
        "fc": fc,
        "sc": sc,
        "township": data.get("乡镇") or "",
        "village": data.get("村") or "",
        "pts": len(track),
        "start": ts[0] if ts else "",
        "end": ts[-1] if ts else "",
    }


def _write_tracks_shp(entries, out_path):
    """entries = [(props, coords)] → 单 shapefile（每小班一条 LineString，WGS84）。

    UTF-8 编码（.cpg 随附），ArcGIS 10.1+ / Pro / QGIS 均可正确读中文属性值。
    """
    import geopandas as gpd
    from shapely.geometry import LineString
    rows = [props for props, _ in entries]
    geoms = [LineString(coords) for _, coords in entries]
    gdf = gpd.GeoDataFrame(rows, geometry=geoms, crs="EPSG:4326")
    gdf.to_file(out_path, encoding="utf-8")


def export_tracks_zip(pid, category=None, fmt="gpx"):
    """导出项目小班轨迹，打包 zip 返回。

    fmt:
      gpx（默认）— 每个有轨迹的小班一个 .gpx（tracks/ 目录，文件名与
                   Excel 轨迹列同源）；ArcGIS 用 GPX To Features 转要素。
      kml        — 单个 .kml（每小班一个 Placemark）；Google Earth/奥维
                   直接打开，ArcGIS 用 KML To Layer 转要素。
      shp        — 单个 shapefile（每小班一条 LineString 要素 + 属性表：
                   小班/分类/年度/乡镇/村/点数/起止时间），WGS84，
                   ArcGIS 10 双击直接打开（推荐）。

    KML/SHP 为线要素格式：不足 2 个有效点的轨迹跳过（GPX 不受影响）。

    Args:
        pid: 项目 ID
        category: 可选，仅导出该分类小班的轨迹（admin 分类下载）。
        fmt: gpx / kml / shp。

    Returns:
        (BytesIO(zip), {"fmt": str, "files": n, "tracks": n, "total_points": m})
    """
    import zipfile
    from xml.sax.saxutils import escape as _xml_escape

    fmt = (fmt or "gpx").strip().lower()
    if fmt not in ("gpx", "kml", "shp"):
        raise ValueError(f"不支持的轨迹格式: {fmt}")

    sc_rows = storage.list_project_subcompartment_rows(pid, category=category)
    sc_rows.sort(key=_sc_sort_key)  # 按调查小班号排序，与 Excel 导出一致
    # 收集有轨迹的小班（同名冲突追序号，与 GPX 文件名规则一致）
    items, used = [], {}
    for sc_row in sc_rows:
        track = (storage.get_extras(sc_row["id"]) or {}).get("track") or []
        if not track:
            continue
        base = _sc_file_base(sc_row)
        k = used.get(base, 0) + 1
        used[base] = k
        items.append((sc_row, track, base if k == 1 else f"{base}_{k}"))
    if not items:
        raise ValueError("该项目暂无轨迹数据")

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        if fmt == "gpx":
            files, total_pts = 0, 0
            for sc_row, track, base in items:
                gpx = _track_gpx(
                    track, _xml_escape(sc_row.get("subcompartment_label") or base))
                if not gpx:
                    continue
                zf.writestr(f"tracks/{base}.gpx", gpx)
                files += 1
                total_pts += len(track)
            if files == 0:
                raise ValueError("该项目暂无轨迹数据")
            buf.seek(0)
            return buf, {"fmt": "gpx", "files": files, "tracks": files,
                         "total_points": total_pts}

        # kml / shp：单文件，全部小班合并
        proj = storage.get_project(pid) or {}
        stem = _dl_stem(proj, category)
        entries = []
        for sc_row, track, base in items:
            coords = _track_coords(track)
            if len(coords) < 2:
                continue  # 单点无法构成线要素
            entries.append((sc_row, track, base, coords))
        if not entries:
            raise ValueError("轨迹均不足 2 个有效点，无法生成线要素")
        total_pts = sum(len(c) for *_, c in entries)

        if fmt == "kml":
            kml = _track_kml(stem, [
                (sc_row.get("subcompartment_label") or base,
                 _track_desc(sc_row, track), coords)
                for sc_row, track, base, coords in entries])
            zf.writestr(f"tracks/{stem}.kml", kml)
            buf.seek(0)
            return buf, {"fmt": "kml", "files": 1, "tracks": len(entries),
                         "total_points": total_pts}

        # shp：临时目录写出全部组件文件（.shp/.shx/.dbf/.prj/.cpg）再入 zip
        import os
        import tempfile
        with tempfile.TemporaryDirectory() as td:
            _write_tracks_shp(
                [(_track_shp_props(sc_row, track), coords)
                 for sc_row, track, base, coords in entries],
                os.path.join(td, stem + ".shp"))
            for fn in sorted(os.listdir(td)):
                with open(os.path.join(td, fn), "rb") as f:
                    zf.writestr(f"tracks/{fn}", f.read())
        buf.seek(0)
        return buf, {"fmt": "shp", "files": 1, "tracks": len(entries),
                     "total_points": total_pts}
