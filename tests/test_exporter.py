# -*- coding: utf-8 -*-
"""exporter 自动化测试 — 双模板导出（基本信息 + 样地）。

模拟 GDB 导入后的数据形态：subcompartment_rows 带 category/data_json，
records 通过 upsert_survey_row 写入（含 samples 子数组）。
"""
import io
import json
import sqlite3
import uuid

import pytest
import openpyxl

from survey.core import storage, exporter, schema as S


class TestTemplateFormulas:
    """模板公式质量门禁（tpl-base 示例行 / tpl-samples 全表）。"""

    @pytest.mark.parametrize("tpl_path", [
        exporter._BASE_TEMPLATE, exporter._SAMPLE_TEMPLATE,
    ], ids=["tpl-base", "tpl-samples"])
    def test_no_circular_reference(self, tpl_path):
        """公式不得引用自身单元格（Excel 打开会提示循环引用）。

        2026-08-21 模板曾出现 =IF(AP5>=0.9,N5,AF5)（AF 列引用自身，
        本意"不合格时保留手填值"），已统一改为 =IF(cond,N5,"")。
        """
        import re
        wb = openpyxl.load_workbook(tpl_path)
        for sn in wb.sheetnames:
            ws = wb[sn]
            for row in ws.iter_rows():
                for c in row:
                    v = c.value
                    if isinstance(v, str) and v.startswith("="):
                        pat = rf"(?<![A-Z$]){re.escape(c.coordinate)}(?!\d)"
                        assert not re.search(pat, v), \
                            f"{tpl_path.name}[{sn}]!{c.coordinate} 公式自引用: {v}"
        wb.close()


@pytest.fixture(autouse=True)
def setup_db(tmp_path, monkeypatch):
    """每个测试用独立临时数据库（避免污染开发库 survey.db）。"""
    monkeypatch.setattr(storage, "_DB_PATH", tmp_path / "test.db")
    storage.init_db()


def _insert_sc_row(pid, data, forest_compartment=0, subcompartment=1,
                   category="人工造林", row_index=0):
    """直接写一条 GDB 形态的小班行（category + data_json）。

    需先建 gdb_files 记录（list_project_subcompartment_rows 按 JOIN 过滤）。
    """
    rid = uuid.uuid4().hex[:12]
    conn = storage._connect()
    try:
        gid = "gdb-test"
        exists = conn.execute("SELECT 1 FROM gdb_files WHERE id=?", (gid,)).fetchone()
        if not exists:
            conn.execute(
                """INSERT INTO gdb_files
                   (id, project_id, file_name, file_hash, layers_json,
                    uploaded_by, uploaded_at)
                   VALUES (?,?,?,?,?,?,?)""",
                (gid, pid, "test.gdb", "hash", "[]", "tester", "2026-08-20 00:00:00"),
            )
        conn.execute(
            """INSERT INTO subcompartment_rows
               (id, batch_id, row_index, data_json, township, village,
                forest_compartment, subcompartment, subcompartment_label,
                tending_area, gdb_id, project_name, category)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (rid, "batch-none", row_index, json.dumps(data, ensure_ascii=False),
             data.get("乡镇", ""), data.get("村", ""),
             forest_compartment, subcompartment,
             f"{forest_compartment}-{subcompartment}", 0,
             gid, "测试项目", category),
        )
        conn.commit()
    finally:
        conn.close()
    return rid


@pytest.fixture
def sample_project():
    """创建测试项目。"""
    return storage.create_project("测试项目", "测试人", "测试乡镇")


@pytest.fixture
def sample_data(sample_project):
    """构造测试数据：2 个人工造林小班，1 条录入记录（含样地）。"""
    pid = sample_project["id"]
    sc1 = _insert_sc_row(pid, {
        "市": "测试市", "县": "测试县", "乡镇": "测试乡", "村": "测试村",
        "林班": 1, "小班": 5, "调查小班号": 5,
        "验收类别": "省级验收", "项目名称": "测试项目",
        "计划年度": 2023, "作业年度": 2023, "林地所有权": "国有",
        "造林树种": "云杉", "上报面积": 100.5, "小班设计株树": 11000,
    }, forest_compartment=1, subcompartment=5, row_index=1)
    sc2 = _insert_sc_row(pid, {
        "市": "测试市", "县": "测试县", "乡镇": "测试乡", "村": "测试村",
        "林班": 1, "小班": 3, "调查小班号": 3,
        "造林树种": "华山松", "上报面积": 80, "需苗量": 9000,
    }, forest_compartment=1, subcompartment=3, row_index=0)

    # sc1 录入：样地 2 个（种植 100+100，成活 90+85）+ 汇总手写项
    # survival_pass 存比率 0-1（新口径，2026-08-21：率类不 ×100）
    storage.upsert_survey_row(pid, "table1", sc1, {
        "survival_pass": "0.85",
        "remark": "测试备注",
        "inspect_time": "2026-08-10",
        "samples": [
            {"no": 1, "area": 100, "planted": 100, "alive": 90,
             "x": 102.123456, "y": 25.654321, "remark": "坡上部"},
            {"no": 2, "area": 100, "planted": 100, "alive": 85, "x": 102.2, "y": 25.6},
        ],
        "sm_grid_area": "5000",
        "sm_grid_count": "4",
        "sm_pole": "完好",
        "sm_film": "无膜",
        "sm_inspector": "李四",
        "sm_inspect_date": "2026-08-21",
        "sm_remark": "汇总备注测试",
    }, "张三")
    return {"pid": pid, "sc1": sc1, "sc2": sc2}


class TestSampleStats:
    """查数株数 = 调查总株数（B34 口径）：round(Σ种植÷个数÷150×网格面积×网格数量)。"""

    BASE = {"samples": [
        {"area": 100, "planted": 100, "alive": 90},
        {"area": 100, "planted": 100, "alive": 85},
    ]}

    def test_stats_formula(self):
        """网格 5000×4、实际个数 2：查数=13333，率=0.875（比率），合格=11666。"""
        stats = exporter._sample_stats(
            dict(self.BASE, sm_grid_area="5000", sm_grid_count="4"))
        assert stats["planted_total"] == 13333
        assert stats["qualified_rate"] == 0.875
        assert stats["qualified_count"] == 11666

    def test_stats_manual_count(self):
        """手写个数 sm_total_count=5 优先：查数=5333，合格=4666。"""
        stats = exporter._sample_stats(
            dict(self.BASE, sm_grid_area="5000", sm_grid_count="4",
                 sm_total_count="5"))
        assert stats["planted_total"] == 5333
        assert stats["qualified_count"] == 4666

    def test_stats_no_grid(self):
        """网格未填（同模板 B32/B33 空参与乘法）：查数=0，合格=0，率不变。"""
        stats = exporter._sample_stats(dict(self.BASE))
        assert stats["planted_total"] == 0
        assert stats["qualified_rate"] == 0.875
        assert stats["qualified_count"] == 0

    def test_stats_empty(self):
        stats = exporter._sample_stats({})
        assert stats["planted_total"] is None
        assert stats["qualified_rate"] is None
        assert stats["qualified_count"] is None

    def test_stats_invalid_samples_excluded(self):
        """无效样地（缺面积或种植株数）不参与统计（2026-08-25 与导出同口径）。"""
        stats = exporter._sample_stats({"samples": [
            {"area": 100, "planted": 100, "alive": 90},
            {"area": 100, "planted": "", "alive": 50},   # 缺种植株数
            {"planted": 100, "alive": 50},               # 缺面积
            {"area": 100, "alive": 50},                  # 缺种植株数
            "垃圾行",
        ]})
        # 仅第 1 个有效：Σ种植=100、Σ成活=90、个数=1
        assert stats["qualified_rate"] == 0.9

    def test_sheet_year_dynamic(self, sample_project):
        """sheet 名年度随项目名：2024 年度项目导出「2024年度人工造林」。"""
        pid = storage.create_project("测试项目(2024 年度)", "测试人", "测试乡镇")["id"]
        _insert_sc_row(pid, {"林班": 1, "小班": 5, "调查小班号": 5, "计划年度": 2023},
                       forest_compartment=1, subcompartment=5)
        output, _ = exporter.export_base(pid)
        wb = openpyxl.load_workbook(output)
        assert "2024年度人工造林" in wb.sheetnames
        assert "2023年度人工造林" not in wb.sheetnames


class TestExportBase:
    def test_returns_bytesio(self, sample_data):
        output, stats = exporter.export_base(sample_data["pid"])
        assert isinstance(output, io.BytesIO)

    def test_stats(self, sample_data):
        _, stats = exporter.export_base(sample_data["pid"])
        assert stats["project"] == "测试项目"
        assert stats["sheets"]["人工造林"] == 2

    def test_three_sheets(self, sample_data):
        output, _ = exporter.export_base(sample_data["pid"])
        wb = openpyxl.load_workbook(output)
        assert len(wb.sheetnames) == 3

    def test_prefilled_green_cols_filled(self, sample_data):
        """绿色预填列从 GDB 数据填充（表1：B调查小班号/M造林树种/AQ设计株树）。"""
        output, _ = exporter.export_base(sample_data["pid"])
        wb = openpyxl.load_workbook(output)
        ws = wb["2023年度人工造林"]
        # 排序（林班,小班）：sc2(小班3) 在前 → 行5；sc1(小班5) → 行6
        assert ws.cell(row=5, column=2).value == 3       # B 调查小班号
        assert ws.cell(row=5, column=13).value == "华山松"  # M 造林树种
        assert ws.cell(row=5, column=43).value == 9000     # AQ 小班设计株树
        assert ws.cell(row=6, column=2).value == 5
        assert ws.cell(row=6, column=43).value == 11000

    def test_input_and_sample_stats(self, sample_data):
        """白色录入列 + 样地统计列（AN/AO/AP）——率类比率 + 0.00% 格式。"""
        output, _ = exporter.export_base(sample_data["pid"])
        wb = openpyxl.load_workbook(output)
        ws = wb["2023年度人工造林"]
        # sc1 在行6：O 成活率分派列 store:false —— 录入值不导出，模板公式代入
        # （AP6=0.875 <0.9 → Excel 打开显示空白）
        assert ws.cell(row=6, column=15).value == '=IF(AP6>=0.9,AP6,"")'
        assert ws.cell(row=6, column=15).number_format == '0.00%'
        # AN 查数株数 = 调查总株数 13333（B34 口径），AO 合格株树 11666，AP 合格率 0.875
        assert ws.cell(row=6, column=40).value == 13333
        assert ws.cell(row=6, column=41).value == 11666
        assert ws.cell(row=6, column=42).value == 0.875
        # 行6 为模板行5 之后的导出行：率类单元格须补设 0.00% 格式（显示 87.50%）
        assert ws.cell(row=6, column=42).number_format == '0.00%'
        # AU 备注（验收人员列插入后右移）
        assert ws.cell(row=6, column=47).value == "测试备注"

    def test_derived_dispatch_columns_formula_only(self, sample_data):
        """store:false 分派列（成活率等级 O/P/Q + 面积 S/AF/AH/AL）一律模板公式：
        即使 data_json 残留旧值也不导出，由 Excel 现算（合格率≥0.9 → 上报面积）。
        """
        pid = sample_data["pid"]
        sc1 = sample_data["sc1"]
        # 残留值模拟（store:false 字段正常情况下不会落库）
        rec = storage.get_survey_rows(pid, "table1")
        base = next((r["data"] or {} for r in rec if r["subcompartment_id"] == sc1), {})
        base.update({
            "survival_replant": "0.875",   # P 列残留
            "verified_pass": "999.9",      # S 列残留
            "construction_area": "888.8",  # AF 列残留
        })
        storage.upsert_survey_row(pid, "table1", sc1, base, "张三",
                                  base_version=(next(r["version"] for r in rec
                                                     if r["subcompartment_id"] == sc1)))
        output, _ = exporter.export_base(pid)
        wb = openpyxl.load_workbook(output)
        ws = wb["2023年度人工造林"]
        # sc1 行6：残留值全部忽略，公式代入（AP6=0.875 在 0.4~0.9 → 待补植列显示率）
        assert ws.cell(row=6, column=15).value == '=IF(AP6>=0.9,AP6,"")'    # O 合格
        assert ws.cell(row=6, column=16).value == '=IF(AND(AP6<0.9,AP6>0.4),AP6,"")'  # P 待补植
        assert ws.cell(row=6, column=17).value == '=IF(AP6<=0.4,AP6,"")'    # Q 失败
        # 面积分派：合格率≥0.9 才回填上报面积 N 列 —— 同样公式承担
        assert ws.cell(row=6, column=19).value == '=IF(AP6>=0.9,N6,"")'     # S 合格面积
        assert ws.cell(row=6, column=32).value == '=IF(AP6>=0.9,N6,"")'     # AF 施工面积
        assert ws.cell(row=6, column=34).value == '=IF(AP6>=0.9,N6,"")'     # AH 建档面积
        assert ws.cell(row=6, column=38).value == '=IF(AP6>=0.9,N6,"")'     # AL 抚育面积

    def test_unsurveyed_row_leaves_input_blank(self, sample_data):
        """未录入小班（sc2 行5）：纯数据列留空，模板公式列代入公式。"""
        output, _ = exporter.export_base(sample_data["pid"])
        wb = openpyxl.load_workbook(output)
        ws = wb["2023年度人工造林"]
        # AN（查数株数）无模板公式 → 留空
        assert ws.cell(row=5, column=40).value in (None, "")
        # O 列模板公式（成活率分派）代入数据行（行5 偏移 0）
        assert ws.cell(row=5, column=15).value == '=IF(AP5>=0.9,AP5,"")'

    def test_percent_ratio_precision(self, sample_data):
        """率类比率保留 4 位小数：0.9524 不被 _fmt_num 的 2 位舍入截成 0.95。"""
        pid = sample_data["pid"]
        sc2 = sample_data["sc2"]
        storage.upsert_survey_row(pid, "table1", sc2, {
            "survival_pass": "0.9524",     # O 列（store:false 残留 → 导出忽略，公式承担）
            "construction_rate": "0.9844",  # AG 列（施工率，手输 percent 落库）
        }, "张三")
        output, _ = exporter.export_base(pid)
        wb = openpyxl.load_workbook(output)
        ws = wb["2023年度人工造林"]
        # sc2 排序后行5：O 分派列公式代入（Excel 打开现算），残留值 0.9524 不导出
        assert ws.cell(row=5, column=15).value == '=IF(AP5>=0.9,AP5,"")'
        assert ws.cell(row=5, column=15).number_format == '0.00%'
        assert ws.cell(row=5, column=33).value == 0.9844   # AG
        assert ws.cell(row=5, column=33).number_format == '0.00%'

    def test_category_filter(self, sample_data):
        """按分类导出：仅保留当前分类 sheet，其余分类 sheet 移除。"""
        output, stats = exporter.export_base(sample_data["pid"], category="人工造林")
        wb = openpyxl.load_workbook(output)
        assert wb.sheetnames == ["2023年度人工造林"]
        assert stats["sheets"] == {"人工造林": 2}
        ws = wb["2023年度人工造林"]
        assert ws.cell(row=5, column=2).value == 3   # 数据仍按序填充

    def test_category_filter_invalid(self, sample_data):
        """未知分类报错（端点转 404）。"""
        with pytest.raises(ValueError, match="未知分类"):
            exporter.export_base(sample_data["pid"], category="水利水保")

    def test_int_like_no_trailing_dot_zero(self, sample_data):
        """小班/年度/株数不可能有小数："2023.0"/"11000.0" → 整数 2023/11000。"""
        pid = sample_data["pid"]
        _insert_sc_row(pid, {
            "市": "测试市", "县": "测试县", "乡镇": "测试乡", "村": "测试村",
            "林班": 1, "小班": 7, "调查小班号": 7,
            "计划年度": "2023.0", "作业年度": "2023.0", "需苗量": "11000.0",
        }, forest_compartment=1, subcompartment=7, row_index=9)
        output, _ = exporter.export_base(pid)
        wb = openpyxl.load_workbook(output)
        ws = wb["2023年度人工造林"]
        # 排序（林班,小班）：3→行5，5→行6，7→行7
        assert ws.cell(row=7, column=10).value == 2023       # J 计划年度
        assert not isinstance(ws.cell(row=7, column=10).value, str)
        assert ws.cell(row=7, column=11).value == 2023       # K 作业年度
        assert ws.cell(row=7, column=43).value == 11000      # AQ 株树

    def test_orig_subcompartment_chinese(self, sample_data):
        """旧数据（含「小班原始」键）：汉字原样保留（G 列，兼容读取）。"""
        pid = sample_data["pid"]
        _insert_sc_row(pid, {
            "市": "测试市", "县": "测试县", "乡镇": "测试乡", "村": "测试村",
            "林班": 1, "小班": 9, "调查小班号": 9, "小班原始": "红9",
        }, forest_compartment=1, subcompartment=9, row_index=9)
        output, _ = exporter.export_base(pid)
        wb = openpyxl.load_workbook(output)
        ws = wb["2023年度人工造林"]
        assert ws.cell(row=7, column=7).value == "红9"   # G 小班

    def test_subcompartment_chinese_new_import(self, sample_data):
        """新导入形态（无「小班原始」）：小班=GDB 原值含汉字（G 列原样），
        调查小班号=数字业务键（B 列）。"""
        pid = sample_data["pid"]
        _insert_sc_row(pid, {
            "市": "测试市", "县": "测试县", "乡镇": "测试乡", "村": "测试村",
            "林班": 1, "小班": "红9", "调查小班号": 9,
        }, forest_compartment=1, subcompartment=9, row_index=9)
        output, _ = exporter.export_base(pid)
        wb = openpyxl.load_workbook(output)
        ws = wb["2023年度人工造林"]
        assert ws.cell(row=7, column=2).value == 9        # B 调查小班号（数字）
        assert ws.cell(row=7, column=7).value == "红9"   # G 小班（汉字原样）

    def test_subcompartment_mapping(self):
        """字段映射：小班→subcompartment_orig（原值），调查小班号→subcompartment（uint）；
        旧键「小班原始」兼容优先。"""
        pf = S.map_subcompartment_to_prefilled(
            {"小班": "红9", "调查小班号": 9})
        assert pf["subcompartment_orig"] == "红9"
        assert pf["subcompartment"] == 9

        # 旧数据：小班被旧导入覆写为数字，小班原始保留原值 → 原值优先
        pf_old = S.map_subcompartment_to_prefilled(
            {"小班": 5, "调查小班号": "5.0", "小班原始": "71"})
        assert pf_old["subcompartment_orig"] == "71"
        assert pf_old["subcompartment"] == 5

        # 无小班字段：subcompartment_orig 缺省，导出回退调查小班号
        pf_min = S.map_subcompartment_to_prefilled({"调查小班号": "12.0"})
        assert pf_min["subcompartment"] == 12
        assert "subcompartment_orig" not in pf_min

    def test_derive_uint_with_digits(self):
        """调查小班号推导：数值直转；含汉字提取数字；无数字→0。"""
        f = storage._derive_uint_with_digits
        assert f("5.0") == 5
        assert f(7) == 7
        assert f("红9") == 9
        assert f("新增17") == 17
        assert f("红") == 0
        assert f("") == 0
        assert f(None) == 0

    def test_checkin_coord_precision(self, sample_data):
        """打卡坐标全精度（GPS 6 位小数），不被 _fmt_num 舍入到 2 位。"""
        pid = sample_data["pid"]
        storage.upsert_survey_row(pid, "table1", sample_data["sc2"], {
            "sample_coord_x": "102.123456", "sample_coord_y": "25.654321",
        }, "李四")
        output, _ = exporter.export_base(pid)
        wb = openpyxl.load_workbook(output)
        ws = wb["2023年度人工造林"]
        assert ws.cell(row=5, column=53).value == 102.123456   # BA 打卡坐标x（验收人员列插入后右移）
        assert ws.cell(row=5, column=54).value == 25.654321    # BB 打卡坐标y

    def test_mgmt_enum_defaults_exported(self, sample_data):
        """管理情况 5 项（Z-AD）空值导出 schema 默认「有」；验收人员/签字留空。"""
        output, _ = exporter.export_base(sample_data["pid"])
        wb = openpyxl.load_workbook(output)
        ws = wb["2023年度人工造林"]
        # sc1 未录入任何管理情况字段 → 全部默认「有」
        for col in range(26, 31):  # Z=26 ... AD=30
            assert ws.cell(row=5, column=col).value == "有", f"col {col}"
        # AR 验收人员文本 / AS 签字 / AV 配合签字 未录入 → 留空
        assert ws.cell(row=5, column=44).value in (None, "")  # AR 验收人员
        assert ws.cell(row=5, column=45).value in (None, "")  # AS 签字
        assert ws.cell(row=5, column=48).value in (None, "")  # AV 配合签字

    def test_mgmt_enum_saved_value_wins(self, sample_data):
        """已录入值优先于默认：mgmt_design=「无」导出「无」。"""
        pid = sample_data["pid"]
        # sc2=小班3 → 排序后行5
        storage.upsert_survey_row(pid, "table1", sample_data["sc2"], {
            "mgmt_design": "无", "mgmt_meeting": "有",
        }, "张三")
        output, _ = exporter.export_base(pid)
        wb = openpyxl.load_workbook(output)
        ws = wb["2023年度人工造林"]
        assert ws.cell(row=5, column=26).value == "无"   # Z 作业设计
        assert ws.cell(row=5, column=27).value == "有"   # AA 会议纪要

    @staticmethod
    def _sign_data_url():
        """构造 120x60 白底黑迹 PNG data URL（模拟签字 canvas）。"""
        import base64 as _b64
        from PIL import Image, ImageDraw
        im = Image.new("RGB", (120, 60), "white")
        d = ImageDraw.Draw(im)
        d.line([(10, 40), (30, 15), (50, 40), (70, 15), (90, 40)], fill="black", width=3)
        buf = io.BytesIO()
        im.save(buf, format="PNG")
        return "data:image/png;base64," + _b64.b64encode(buf.getvalue()).decode()

    def test_sign_images_inserted(self, sample_data):
        """手写签字导出为图片：6 个签字位 AS/AV~AZ（验收人员列插入后右移），
        格内不写文本；AR 验收人员文本列写录入值。"""
        pid = sample_data["pid"]
        url = self._sign_data_url()
        storage.upsert_survey_row(pid, "table1", sample_data["sc2"], {
            "inspector": "张三",
            "inspector_sign": url, "co_inspector_sign": url,
            "co_inspector_sign2": url, "co_inspector_sign3": url,
            "co_inspector_sign4": url, "co_inspector_sign5": url,
        }, "张三")
        output, _ = exporter.export_base(pid)
        wb = openpyxl.load_workbook(output)
        ws = wb["2023年度人工造林"]
        # 行5（小班3）6 列签字 → 6 张图片（AS/AV/AW/AX/AY/AZ）
        assert len(ws._images) == 6
        cells = sorted((img.anchor._from.row + 1, img.anchor._from.col + 1)
                       for img in ws._images)
        assert cells == [(5, 45), (5, 48), (5, 49),
                         (5, 50), (5, 51), (5, 52)]
        # 固定显示范围：宽 ≤200px、高 ≤64px
        for img in ws._images:
            assert img.width <= 200 and img.height <= 64
        # AR 验收人员文本列写录入值；签字格本身不写 data URL 文本
        assert ws.cell(row=5, column=44).value == "张三"   # AR 验收人员
        for col in (45, 48, 49, 50, 51, 52):
            assert ws.cell(row=5, column=col).value in (None, "")

    def test_sign_blank_skipped(self, sample_data):
        """无签字 → 不插图片、单元格留空。"""
        output, _ = exporter.export_base(sample_data["pid"])
        wb = openpyxl.load_workbook(output)
        for sn in wb.sheetnames:
            assert len(wb[sn]._images) == 0


class TestExportSamples:
    def test_returns_bytesio(self, sample_data):
        output, stats = exporter.export_samples(sample_data["pid"])
        assert isinstance(output, io.BytesIO)
        assert stats["project"] == "测试项目"

    def test_block_structure(self, sample_data):
        """每小班一个 39 行块：sc2 块1（行1-39），sc1 块2（行40-78）。"""
        output, _ = exporter.export_samples(sample_data["pid"])
        wb = openpyxl.load_workbook(output)
        assert "人工造林" in wb.sheetnames
        ws = wb["人工造林"]
        # 块2 起始行 40 的标题 + 块内样地数据（行 43-44，数据起始 40+3）
        assert "样地调查表" in str(ws.cell(row=40, column=1).value)
        # 块1 的 R39 为备注行（模板自带标签，块2 覆盖为标题）
        assert ws.cell(row=39, column=1).value == "备注"
        # sc1 样地1：行 40+3=43，A=样地号1，C=种植100，D=成活90
        assert ws.cell(row=43, column=1).value == 1
        assert ws.cell(row=43, column=3).value == 100
        assert ws.cell(row=43, column=4).value == 90
        # 死亡株数公式（E 列）
        assert ws.cell(row=43, column=5).value == "=C43-D43"
        # 样地备注（I 列）；样地2 无备注 → 空
        assert ws.cell(row=43, column=9).value == "坡上部"
        assert ws.cell(row=44, column=9).value in (None, "")

    def test_summary_fields(self, sample_data):
        """汇总区 R27-R39：个数/手写项写入，公式带除0守卫并按块偏移。"""
        output, _ = exporter.export_samples(
            sample_data["pid"], subcompartment_id=sample_data["sc1"])
        wb = openpyxl.load_workbook(output)
        ws = wb["人工造林-5"]
        # B27 总样地个数 = 样地数（代码计算写入）
        assert ws.cell(row=27, column=2).value == 2
        # B28-30 SUM 公式保留；B31/B34 带 IF 除0守卫
        assert ws.cell(row=28, column=2).value == "=SUM(B4:B26)"
        assert ws.cell(row=31, column=2).value == '=IF(B29=0,"",B30/B29)'
        assert ws.cell(row=34, column=2).value == \
            '=IF(OR(B27=0,B29=0),"",ROUND(B29/B27/150*B32*B33,0))'
        # 手写录入项（sm_* → B32-B39）
        assert ws.cell(row=32, column=2).value == 5000.0
        assert ws.cell(row=33, column=2).value == 4.0
        assert ws.cell(row=35, column=2).value == "完好"
        assert ws.cell(row=36, column=2).value == "无膜"
        assert ws.cell(row=37, column=2).value == "李四"
        assert ws.cell(row=38, column=2).value == "2026-08-21"
        assert ws.cell(row=39, column=2).value == "汇总备注测试"

    def test_total_count_manual_override(self, sample_data):
        """B27 总样地个数：手写 sm_total_count 优先（>0 生效），无效/空回退样地数。"""
        pid = sample_data["pid"]
        sc1 = sample_data["sc1"]
        # 手写覆盖：个数写 5（≠ 实际样地数 2）
        rec = storage.get_survey_rows(pid, "table1")
        d = next(r["data"] for r in rec if r["subcompartment_id"] == sc1)
        d["sm_total_count"] = "5"
        storage.upsert_survey_row(pid, "table1", sc1, d, "张三")
        output, _ = exporter.export_samples(pid, subcompartment_id=sc1)
        wb = openpyxl.load_workbook(output)
        assert wb["人工造林-5"].cell(row=27, column=2).value == 5

        # 无效值（0/非数字）回退实际样地数
        for bad in ("0", "-1", "abc", ""):
            d["sm_total_count"] = bad
            storage.upsert_survey_row(pid, "table1", sc1, d, "张三")
            output, _ = exporter.export_samples(pid, subcompartment_id=sc1)
            wb = openpyxl.load_workbook(output)
            assert wb["人工造林-5"].cell(row=27, column=2).value == 2, bad

    def test_summary_block2_shift(self, sample_data):
        """块2 汇总公式行号偏移 + 手写项按块写入（块2=sc1 有录入）。"""
        output, _ = exporter.export_samples(sample_data["pid"])
        wb = openpyxl.load_workbook(output)
        ws = wb["人工造林"]
        # 排序后 sc2(小班3)=块1、sc1(小班5)=块2（行40起）
        # 块2 B31 公式偏移为 =IF(B68=0,"",B69/B68)
        assert ws.cell(row=40 + 30, column=2).value == '=IF(B68=0,"",B69/B68)'
        # 块2 = sc1：总样地个数 2、网格面积/备注照常写入
        assert ws.cell(row=40 + 26, column=2).value == 2
        assert ws.cell(row=40 + 31, column=2).value == 5000.0
        assert ws.cell(row=40 + 38, column=2).value == "汇总备注测试"

    def test_empty_category_skipped(self, sample_data):
        """无数据的分类跳过（不建 sheet，不崩）。"""
        output, stats = exporter.export_samples(sample_data["pid"])
        assert "封山育林" not in stats["sheets"]
        wb = openpyxl.load_workbook(output)
        assert "封山育林" not in wb.sheetnames
        assert wb.sheetnames == ["人工造林"]

    def test_single_subcompartment(self, sample_data):
        """单小班模式：仅当前小班当前分类，sheet 名「分类-调查小班号」，只有一个块。"""
        output, stats = exporter.export_samples(
            sample_data["pid"], subcompartment_id=sample_data["sc1"])
        wb = openpyxl.load_workbook(output)
        assert wb.sheetnames == ["人工造林-5"]
        ws = wb["人工造林-5"]
        # 只有块1：sc1 样地1 在行 4（块1数据起始），行 40（块2标题位）应为空
        assert ws.cell(row=4, column=1).value == 1
        assert ws.cell(row=4, column=3).value == 100
        assert ws.cell(row=4, column=4).value == 90
        assert ws.cell(row=40, column=1).value is None
        assert stats["sheets"] == {"人工造林-5": 1}

    def test_title_and_coord_precision(self, sample_data):
        """R1 标题含项目类型+调查小班号；R2 年度县乡；坐标写全精度浮点。"""
        output, _ = exporter.export_samples(
            sample_data["pid"], subcompartment_id=sample_data["sc1"])
        wb = openpyxl.load_workbook(output)
        ws = wb["人工造林-5"]
        # R1 标题：项目名称+样地调查表（类型）+ 调查小班号
        r1 = str(ws.cell(row=1, column=1).value)
        assert "样地调查表（人工造林）" in r1
        assert "调查小班号5" in r1
        # R2 年度县乡（A2:E2 合并区）
        r2 = str(ws.cell(row=2, column=1).value)
        assert "年度" in r2
        # 坐标全精度：x=102.123456 不能被舍成 102.12（数据起始 R4）
        assert ws.cell(row=4, column=6).value == 102.123456
        assert ws.cell(row=4, column=7).value == 25.654321

    def test_single_subcompartment_not_found(self, sample_data):
        """单小班模式：小班不存在或不属于该项目时 404 语义（ValueError）。"""
        with pytest.raises(ValueError, match="小班不存在"):
            exporter.export_samples(sample_data["pid"], subcompartment_id="no_such_id")

    def test_category_filter(self, sample_data):
        """分类过滤（admin 分类下载）：仅导出该分类一个 sheet，每分类一个文件。"""
        output, stats = exporter.export_samples(sample_data["pid"], category="人工造林")
        wb = openpyxl.load_workbook(output)
        assert wb.sheetnames == ["人工造林"]
        assert stats["sheets"] == {"人工造林": 2}

    def test_category_filter_empty_raises(self, sample_data):
        """分类过滤：该分类无小班时 ValueError（端点转 404）。"""
        with pytest.raises(ValueError, match="暂无小班数据"):
            exporter.export_samples(sample_data["pid"], category="封山育林")

    def test_samples_zip_per_subcompartment(self, sample_data):
        """分类下载样地 zip：仅含有效样地的小班出文件；
        文件名 {调查小班号}号调查小班-{分类}-{年度}.xlsx（无林班前缀，2026-08-25）。"""
        import zipfile
        buf, stats = exporter.export_samples_zip(sample_data["pid"], category="人工造林")
        # sc1（小班5，2 个有效样地）出文件；sc2（小班3，无录入记录）跳过
        assert stats["files"] == 1
        with zipfile.ZipFile(buf) as zf:
            names = sorted(zf.namelist())
        assert names == ["5号调查小班-人工造林-2023.xlsx"]
        # 每个 xlsx 可打开且为单小班模式（一个 sheet 一个块）
        with zipfile.ZipFile(buf) as zf:
            wb = openpyxl.load_workbook(io.BytesIO(
                zf.read("5号调查小班-人工造林-2023.xlsx")))
        assert wb.sheetnames == ["人工造林-5"]
        assert wb["人工造林-5"].cell(row=4, column=3).value == 100  # 样地1 面积

    def test_invalid_sample_rows_filtered(self, sample_data):
        """无效样地行（缺面积/种植株数）不写入块，B27 回退计数只算有效样地。"""
        pid = sample_data["pid"]
        sc1 = sample_data["sc1"]
        rec = storage.get_survey_rows(pid, "table1")
        d = next(r["data"] for r in rec if r["subcompartment_id"] == sc1)
        d["samples"] = [
            {"no": 1, "area": 100, "planted": 100, "alive": 90},
            {"no": 2, "area": 100, "planted": "", "alive": 50},   # 缺种植株数 → 过滤
            {"no": 3, "planted": 80, "alive": 40},                # 缺面积 → 过滤
        ]
        storage.upsert_survey_row(pid, "table1", sc1, d, "张三")
        output, _ = exporter.export_samples(pid, subcompartment_id=sc1)
        wb = openpyxl.load_workbook(output)
        ws = wb["人工造林-5"]
        # 行4 仅样地1；行5（样地2 位置）为空
        assert ws.cell(row=4, column=1).value == 1
        assert ws.cell(row=4, column=3).value == 100
        assert ws.cell(row=5, column=1).value in (None, "")
        # B27 回退计数 = 有效样地数 1
        assert ws.cell(row=27, column=2).value == 1

    def test_samples_zip_empty_category_raises(self, sample_data):
        """样地 zip：该分类无小班时 ValueError（端点转 404）。"""
        with pytest.raises(ValueError, match="暂无小班数据"):
            exporter.export_samples_zip(sample_data["pid"], category="封山育林")

    def test_samples_singlefile(self, sample_data):
        """单文件版本（2026-08-25）：一个 xlsx 每有效样地小班一个 sheet
        「分类-调查小班号」；无有效样地小班不出 sheet；分类过滤生效。"""
        output, stats = exporter.export_samples_singlefile(sample_data["pid"])
        wb = openpyxl.load_workbook(output)
        # 仅 sc1（小班5）有有效样地；sc2 无录入记录
        assert wb.sheetnames == ["人工造林-5"]
        assert stats["sheets"] == {"人工造林-5": 1}
        ws = wb["人工造林-5"]
        assert ws.cell(row=4, column=3).value == 100  # 样地1 面积
        # 分类过滤：封山育林无有效样地 → ValueError
        with pytest.raises(ValueError, match="暂无样地数据"):
            exporter.export_samples_singlefile(sample_data["pid"], category="封山育林")


class TestInspectDateFilter:
    """验收日期过滤（2026-08-26 D28）：按记录 inspect_time 单日过滤，
    base/样地 zip/样地单文件三处同口径。"""

    @pytest.fixture()
    def dated(self, sample_data):
        """给 sc1 的记录加验收时间 2026-08-20（sc2 无记录）。"""
        pid = sample_data["pid"]
        rec = storage.get_survey_rows(pid, "table1")[0]
        d = dict(rec["data"])
        d["inspect_time"] = "2026-08-20"
        storage.upsert_survey_row(pid, "table1", sample_data["sc1"], d, "张三")
        return sample_data

    def test_base_filtered(self, dated):
        """基本信息：匹配日期仅导出该小班；不匹配日期无数据报错。"""
        pid = dated["pid"]
        out, stats = exporter.export_base(pid, inspect_date="2026-08-20")
        assert stats["sheets"]["人工造林"] == 1
        with pytest.raises(ValueError, match="无录入数据"):
            exporter.export_base(pid, inspect_date="2026-08-19")

    def test_base_unfiltered_unchanged(self, dated):
        """不传日期保持原行为：全部小班（含未录入）。"""
        _out, stats = exporter.export_base(dated["pid"])
        assert stats["sheets"]["人工造林"] == 2

    def test_samples_zip_filtered(self, dated):
        pid = dated["pid"]
        buf, st = exporter.export_samples_zip(pid, inspect_date="2026-08-20")
        assert st["files"] == 1
        with pytest.raises(ValueError, match="无样地数据"):
            exporter.export_samples_zip(pid, inspect_date="2026-08-19")

    def test_samples_singlefile_filtered(self, dated):
        out, stats = exporter.export_samples_singlefile(
            dated["pid"], inspect_date="2026-08-20")
        assert list(stats["sheets"]) == ["人工造林-5"]

    def test_cache_bypass_when_filtered(self, dated):
        """过滤导出绕过缓存（不落盘），全量导出不受影响仍走缓存。"""
        from survey.core import export_cache
        pid = dated["pid"]
        out, st = export_cache.cached_or_generate(pid, "人工造林", "samples",
                                                  inspect_date="2026-08-20")
        assert isinstance(out, io.BytesIO) and st["files"] == 1
        # 未生成缓存文件（绕过）
        assert export_cache.load_cached(pid, "人工造林", "samples") is None

    def test_base_county_filter(self, sample_data):
        """县过滤（2026-08-26 D30）：按小班 GDB「县」字段匹配；不匹配报错。"""
        pid = sample_data["pid"]
        # fixture 小班 data_json「县」默认值
        row = storage.get_subcompartment_row(sample_data["sc1"])
        county = (row.get("data") or {}).get("县")
        assert county  # fixture 必须含县字段
        _out, stats = exporter.export_base(pid, county=county)
        assert stats["sheets"]["人工造林"] == 2  # 两小班同县
        with pytest.raises(ValueError, match="无录入数据"):
            exporter.export_base(pid, county="不存在县")

    def test_base_county_and_date_combined(self, dated):
        """县 + 验收日期叠加过滤：sc1 有记录（日期匹配），sc2 无记录。"""
        pid = dated["pid"]
        county = (storage.get_subcompartment_row(dated["sc1"]).get("data") or {}).get("县")
        _out, stats = exporter.export_base(pid, county=county,
                                           inspect_date="2026-08-20")
        assert stats["sheets"]["人工造林"] == 1
        # 日期不匹配 + 县匹配 → 无数据
        with pytest.raises(ValueError, match="无录入数据"):
            exporter.export_base(pid, county=county, inspect_date="2026-08-19")


def _read_shp_from_zip(buf):
    """解包轨迹 zip 并用 geopandas 读回全部 shapefile（2026-08-25 起每小班
    一个独立 .shp，目录隔离 {调查小班号}号调查小班-{分类}-{年度}/），合并返回。"""
    import geopandas as gpd
    import os
    import tempfile
    import zipfile
    with zipfile.ZipFile(buf) as zf:
        names = zf.namelist()
        with tempfile.TemporaryDirectory() as td:
            for n in names:
                zf.extract(n, td)
            shps = sorted(n for n in names if n.endswith(".shp"))
            gdfs = [gpd.read_file(os.path.join(td, s)) for s in shps]
            return gpd.pd.concat(gdfs, ignore_index=True) if len(gdfs) > 1 else gdfs[0]


class TestExportTracks:
    def test_no_tracks_raises(self, sample_data):
        """无轨迹时抛 ValueError（端点转 404）。"""
        with pytest.raises(ValueError, match="暂无轨迹"):
            exporter.export_tracks_zip(sample_data["pid"])

    def test_category_filter(self, sample_data):
        """分类过滤：仅导出该分类小班的轨迹（单 shapefile，geopandas 读回验证）。"""
        pid = sample_data["pid"]
        sc3 = _insert_sc_row(pid, {
            "乡镇": "测试乡", "村": "测试村", "林班": 2, "小班": 7, "调查小班号": 7,
        }, forest_compartment=2, subcompartment=7, category="封山育林", row_index=2)
        storage.save_track(sample_data["sc1"], [
            {"lng": 102.1, "lat": 25.6, "t": "2026-08-21T10:00:00"},
            {"lng": 102.11, "lat": 25.61, "t": "2026-08-21T10:01:00"},
        ])
        storage.save_track(sc3, [
            {"lng": 102.3, "lat": 25.7, "t": "2026-08-21T11:00:00"},
            {"lng": 102.31, "lat": 25.71, "t": "2026-08-21T11:01:00"},
        ])
        # 过滤封山育林 → 仅 sc3 一条要素
        buf, stats = exporter.export_tracks_zip(pid, category="封山育林")
        assert stats["fmt"] == "shp" and stats["tracks"] == 1
        gdf = _read_shp_from_zip(buf)
        assert list(gdf["name"]) == ["2-7"]
        assert list(gdf["category"]) == ["封山育林"]
        # 不过滤 → 两类各一条
        buf_all, stats_all = exporter.export_tracks_zip(pid)
        assert stats_all["tracks"] == 2
        gdf_all = _read_shp_from_zip(buf_all)
        assert sorted(gdf_all["name"]) == ["1-5", "2-7"]
        assert sorted(gdf_all["category"]) == ["人工造林", "封山育林"]
        # 过滤无轨迹分类 → ValueError
        _insert_sc_row(pid, {
            "乡镇": "测试乡", "村": "测试村", "林班": 3, "小班": 9, "调查小班号": 9,
        }, forest_compartment=3, subcompartment=9, category="退化林修复", row_index=3)
        with pytest.raises(ValueError, match="暂无轨迹"):
            exporter.export_tracks_zip(pid, category="退化林修复")

    def test_shp_per_subcompartment(self, sample_data):
        """SHP 按小班分文件（2026-08-25 R17）：每小班独立 shapefile 目录隔离，
        geopandas 读回验证内容与组件齐全。"""
        import zipfile
        pid = sample_data["pid"]
        storage.save_track(sample_data["sc1"], [
            {"lng": 102.1, "lat": 25.6, "t": "2026-08-21T10:00:00"},
            {"lng": 102.11, "lat": 25.61, "t": "2026-08-21T10:01:00"},
            {"lng": 102.12, "lat": 25.62, "t": "2026-08-21T10:02:00"},
        ])
        buf, stats = exporter.export_tracks_zip(pid)
        assert stats["fmt"] == "shp" and stats["tracks"] == 1 and stats["files"] == 1
        with zipfile.ZipFile(buf) as zf:
            names = zf.namelist()
            # 目录隔离：{调查小班号}号调查小班-{分类}-{年度}/ 前缀
            prefix = "5号调查小班-人工造林-2023/"
            assert any(n.startswith(prefix) and n.endswith(".shp") for n in names)
            # shapefile 组件齐全（.shp/.shx/.dbf/.prj/.cpg）
            for ext in (".shp", ".shx", ".dbf", ".prj", ".cpg"):
                assert any(n.startswith(prefix) and n.endswith(ext) for n in names), ext
        gdf = _read_shp_from_zip(buf)
        assert len(gdf) == 1
        row = gdf.iloc[0]
        assert row.geometry.geom_type == "LineString"
        assert round(row.geometry.coords[0][0], 4) == 102.1
        assert row["category"] == "人工造林"
        assert row["township"] == "测试乡" and row["village"] == "测试村"
        assert row["pts"] == 3
        assert str(row["start"]).startswith("2026-08-21")
        assert gdf.crs is not None and gdf.crs.to_epsg() == 4326  # WGS84

    def test_buf_position_at_start(self, sample_data):
        """回归：返回的 BytesIO 必须在位置 0（2026-08-24 生产空下载事故）。

        ZipFile.close() 写完中央目录后指针停在 EOF；seek(0) 若写在 with
        块内会被 close 推回末尾 → send_file 从当前位置读出 0 字节
        （200 + Content-Length 正确但空 body）。测试里用 ZipFile 打开
        不受指针影响，唯本断言能兜住。
        """
        storage.save_track(sample_data["sc1"], [
            {"lng": 102.1, "lat": 25.6, "t": "2026-08-21T10:00:00"},
            {"lng": 102.11, "lat": 25.61, "t": "2026-08-21T10:01:00"},
        ])
        buf, _stats = exporter.export_tracks_zip(sample_data["pid"])
        assert buf.tell() == 0
        data = buf.read()  # 模拟 send_file：从当前位置读到尾
        assert data[:4] == b"PK\x03\x04" and len(data) > 100

    def test_dedupe_consecutive_points(self, sample_data):
        """连续重复点位去重：GPS 精度受限常回传相同坐标（生产数据实测）。"""
        storage.save_track(sample_data["sc1"], [
            {"lng": 102.1, "lat": 25.6, "t": "2026-08-21T10:00:00"},
            {"lng": 102.1, "lat": 25.6, "t": "2026-08-21T10:00:05"},   # 重复
            {"lng": 102.1, "lat": 25.6, "t": "2026-08-21T10:00:10"},   # 重复
            {"lng": 102.11, "lat": 25.61, "t": "2026-08-21T10:01:00"},
            {"lng": 102.11, "lat": 25.61, "t": "2026-08-21T10:01:05"},  # 重复
        ])
        buf, stats = exporter.export_tracks_zip(sample_data["pid"])
        gdf = _read_shp_from_zip(buf)
        assert stats["total_points"] == 2
        assert gdf.iloc[0]["pts"] == 2
        assert len(gdf.iloc[0].geometry.coords) == 2

    def test_all_single_point_raises(self, sample_data):
        """全部轨迹不足 2 点 → ValueError（线要素无法生成）。"""
        storage.save_track(sample_data["sc1"], [
            {"lng": 102.1, "lat": 25.6, "t": "2026-08-21T10:00:00"},
        ])
        with pytest.raises(ValueError, match="不足 2 个有效点"):
            exporter.export_tracks_zip(sample_data["pid"])


class TestExportCache:
    """导出缓存层（2026-08-25 D25）：指纹失效 / 命中复用 / 项目级合并 / prefetch。"""

    @pytest.fixture(autouse=True)
    def _cleanup_cache(self):
        import shutil
        from survey.core import export_cache
        yield
        shutil.rmtree(export_cache.cache_root(), ignore_errors=True)

    def test_cache_hit_and_invalidate(self, sample_data):
        """未命中生成并回写；命中复用（文件不变）；数据变更后指纹失效重生成。"""
        from survey.core import export_cache
        pid = sample_data["pid"]
        path1, stats1 = export_cache.cached_or_generate(pid, "人工造林", "samples")
        assert path1.exists() and stats1["files"] == 1
        mtime1 = path1.stat().st_mtime_ns
        # 二次：命中 → 同一磁盘文件，未重生成
        path2, stats2 = export_cache.cached_or_generate(pid, "人工造林", "samples")
        assert path2 == path1
        assert path2.stat().st_mtime_ns == mtime1
        assert stats2 == stats1
        # 数据变更（records.updated_at 变）→ 指纹失效 → 重写缓存
        rec = storage.get_survey_rows(pid, "table1")[0]
        d = dict(rec["data"])
        d["remark"] = "改一下触发指纹变化"
        storage.upsert_survey_row(pid, "table1", sample_data["sc1"], d, "张三")
        assert export_cache.load_cached(pid, "人工造林", "samples") is None
        path3, _ = export_cache.cached_or_generate(pid, "人工造林", "samples")
        assert path3.stat().st_mtime_ns != mtime1

    def test_track_fingerprint_follows_extras(self, sample_data):
        """轨迹保存（extras.updated_at 变）也使指纹失效。"""
        from survey.core import export_cache
        pid = sample_data["pid"]
        export_cache.cached_or_generate(pid, "人工造林", "samples")
        fp1 = export_cache.fingerprint(pid, "人工造林")
        storage.save_track(sample_data["sc1"], [
            {"lng": 102.1, "lat": 25.6, "t": "2026-08-21T10:00:00"},
            {"lng": 102.11, "lat": 25.61, "t": "2026-08-21T10:01:00"},
        ])
        assert export_cache.fingerprint(pid, "人工造林") != fp1

    def test_project_merge_zip(self, sample_data):
        """项目级合并 zip：各分类缓存文件拼进一个 zip（BytesIO 位置 0）。"""
        import zipfile
        from survey.core import export_cache
        buf = export_cache.cached_or_generate_project(sample_data["pid"], "samples")
        assert buf.tell() == 0
        with zipfile.ZipFile(buf) as zf:
            names = zf.namelist()
        assert names == ["5号调查小班-人工造林-2023.xlsx"]

    def test_project_merge_all_empty_raises(self, sample_project):
        """项目级合并：全部分类无数据 → ValueError。"""
        from survey.core import export_cache
        with pytest.raises(ValueError, match="暂无样地数据"):
            export_cache.cached_or_generate_project(sample_project["id"], "samples")

    def test_prefetch_quiet_and_generate(self, sample_data, monkeypatch):
        """prefetch：数据刚更新（静默窗口内）跳过；窗口外生成；再跑命中跳过。"""
        from survey.core import export_cache
        pid = sample_data["pid"]
        proj = storage.get_project(pid)
        monkeypatch.setattr(export_cache.storage, "list_projects", lambda: [proj])
        # 静默窗口内（fixture 刚写 records）→ 跳过不生成
        s = export_cache.prefetch(quiet_minutes=10)
        assert s["skipped_quiet"] >= 1 and not s["generated"]
        assert export_cache.load_cached(pid, "人工造林", "samples") is None
        # 窗口设 0（视为数据已静默）→ 生成；无轨迹的轨迹缓存走 emptied
        s2 = export_cache.prefetch(quiet_minutes=0)
        assert s2["generated"] >= 1
        hit = export_cache.load_cached(pid, "人工造林", "samples")
        assert hit is not None
        # 再跑 → 指纹一致跳过（tracks 无数据仍走 emptied，不算 generated）
        s3 = export_cache.prefetch(quiet_minutes=0)
        assert s3["skipped_fresh"] >= 1 and not s3["generated"]

    def test_cleanup_project(self, sample_data):
        """项目删除 → 缓存目录清理。"""
        from survey.core import export_cache
        pid = sample_data["pid"]
        export_cache.cached_or_generate(pid, "人工造林", "samples")
        assert export_cache.load_cached(pid, "人工造林", "samples") is not None
        export_cache.cleanup_project(pid)
        assert export_cache.load_cached(pid, "人工造林", "samples") is None
