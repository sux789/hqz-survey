# -*- coding: utf-8 -*-
"""导出结果双跑对比 — 改动导出代码后的回归防线（AGENTS.md）。

用法:
    python tools/compare_export.py <目录A> <目录B>     # 对比两个导出目录（按文件名配对）
    python tools/compare_export.py <文件A> <文件B>     # 对比两个 xlsx

逐 sheet、逐单元格对比值（浮点容差 1e-9），并对比每 sheet 图片数。
有差异时打印前 30 处并以退出码 1 结束；全一致打印一致率并退出 0。
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl  # noqa: E402

_TOL = 1e-9
_MAX_DIFFS = 30


def _cells(path):
    """{sheet: {coord: value}} + {sheet: 图片数}。"""
    wb = openpyxl.load_workbook(path)
    data, images = {}, {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        data[sn] = {c.coordinate: c.value for row in ws.iter_rows() for c in row
                    if c.value not in (None, "")}
        images[sn] = len(ws._images)
    wb.close()
    return data, images


def _eq(a, b):
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return abs(float(a) - float(b)) <= _TOL
    return a == b


def compare(file_a, file_b, diffs):
    a, aimg = _cells(file_a)
    b, bimg = _cells(file_b)
    total = 0
    for sn in sorted(set(a) | set(b)):
        if sn not in a:
            diffs.append(f"{Path(file_a).name}:[{sn}] sheet 仅存在于 B")
            continue
        if sn not in b:
            diffs.append(f"{Path(file_a).name}:[{sn}] sheet 仅存在于 A")
            continue
        if aimg[sn] != bimg[sn]:
            diffs.append(f"{Path(file_a).name}:[{sn}] 图片数 A={aimg[sn]} B={bimg[sn]}")
        for coord in sorted(set(a[sn]) | set(b[sn])):
            total += 1
            va, vb = a[sn].get(coord), b[sn].get(coord)
            if not _eq(va, vb):
                diffs.append(f"{Path(file_a).name}:[{sn}]{coord} A={va!r} B={vb!r}")
    return total


def main():
    pa, pb = Path(sys.argv[1]), Path(sys.argv[2])
    diffs = []
    total = 0
    if pa.is_dir() and pb.is_dir():
        files_a = {f.name: f for f in pa.glob("*.xlsx")}
        files_b = {f.name: f for f in pb.glob("*.xlsx")}
        for name in sorted(set(files_a) | set(files_b)):
            if name not in files_a:
                diffs.append(f"{name} 仅存在于 B")
            elif name not in files_b:
                diffs.append(f"{name} 仅存在于 A")
            else:
                total += compare(str(files_a[name]), str(files_b[name]), diffs)
    else:
        total = compare(str(pa), str(pb), diffs)

    if diffs:
        print(f"FAIL: {len(diffs)} 处差异（共对比 {total} 单元格）")
        for d in diffs[:_MAX_DIFFS]:
            print("  " + d)
        if len(diffs) > _MAX_DIFFS:
            print(f"  ... 其余 {len(diffs) - _MAX_DIFFS} 处省略")
        sys.exit(1)
    print(f"PASS: 全部一致（{total} 单元格）")


if __name__ == "__main__":
    main()
