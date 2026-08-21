# -*- coding: utf-8 -*-
"""生产导出验证（一次性脚本）：签 session cookie 调生产导出 API，核对本次改动——
  1. 样地单小班导出：I3=备注表头；B31/B34 除0守卫公式完整；A2/R1 结构
  2. 基本信息导出：表1 小班1 行 AN=860（B34 新口径；旧口径Σ种植=43）/
     AO=820 / AP=95.35
  3. 整项目基本信息导出不 500

生产数据（2023年度 人工造林 小班1，subcompartment_id=97d6a528e39a）：
  samples 5 条：planted 16/14/13/0/0（Σ=43），alive 15/14/12/0/0（Σ=41）；
  sm_grid_area=5000, sm_grid_count=3, sm_total_count=5
  → AN = round(43÷5÷150×5000×3) = 860；AP = round(41/43×100,2) = 95.35；
    AO = round(860×95.35/100) = 820

用法: python tools/verify_prod_export.py <secret_key>
"""
import hashlib
import io
import ssl
import sys
import urllib.request
from urllib.error import HTTPError

from flask import Flask
from flask.sessions import SecureCookieSessionInterface, TaggedJSONSerializer
from itsdangerous import URLSafeTimedSerializer

ADMIN = "https://forest.bibook.top/survey-admin"
USER = "https://forest.bibook.top/survey"
PID = "b0b4d046c479"  # 玉溪市滇中山地石漠化综合治理项目(2023 年度)
SC_ID = "97d6a528e39a"  # 人工造林 小班1（林班0）


def sign_cookie(secret_key: str, user_id: int = 1) -> str:
    """按 flask SecureCookieSessionInterface 口径签 session cookie。"""
    app = Flask(__name__)
    app.secret_key = secret_key
    si = SecureCookieSessionInterface()
    s = URLSafeTimedSerializer(
        secret_key,
        salt=si.salt,  # 'session' = cookie 名
        serializer=TaggedJSONSerializer(),
        signer_kwargs=dict(
            key_derivation=si.key_derivation, digest_method=hashlib.sha1
        ),
    )
    return "session=" + s.dumps({"user_id": user_id, "username": "hqz001"})


def _ssl_ctx():
    """本地 pyenv 常缺 CA 证书：优先 certifi，回退不校验（一次性只读验证）。"""
    try:
        import certifi

        return ssl.create_default_context(cafile=certifi.where())
    except ImportError:
        return ssl._create_unverified_context()


def fetch(url: str, cookie: str, timeout: int = 120) -> bytes:
    req = urllib.request.Request(url, headers={"Cookie": cookie})
    with urllib.request.urlopen(req, timeout=timeout, context=_ssl_ctx()) as resp:
        return resp.read()


def main():
    secret_key = sys.argv[1]
    cookie = sign_cookie(secret_key)
    print("cookie signed ok")
    import openpyxl

    results = []

    # ── 1) 单小班样地导出（用户端，快）──
    blob = fetch(f"{USER}/api/projects/{PID}/export_samples?sc={SC_ID}", cookie)
    print(f"单小班样地导出: {len(blob)} bytes")
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    ws = wb.worksheets[0]
    print(f"  sheet={ws.title!r} A2={ws['A2'].value!r} R1={ws['A1'].value!r}")
    i3 = ws.cell(row=3, column=9).value
    b31 = ws.cell(row=31, column=2).value
    b34 = ws.cell(row=34, column=2).value
    f31 = isinstance(b31, str) and "IF" in b31 and "B29=0" in b31
    f34 = isinstance(b34, str) and "OR(B27=0,B29=0)" in b34 and "ROUND" in b34
    results.append((f"I3=备注表头（实际 {i3!r}）", i3 == "备注"))
    results.append((f"B31 除0守卫（实际 {b31!r}）", bool(f31)))
    results.append((f"B34 除0守卫（实际 {b34!r}）", bool(f34)))
    # 数据行：样地1 planted=16 alive=15 → B/C/D 列 + I 列备注（生产暂无备注数据）
    r4 = [ws.cell(row=4, column=c).value for c in (2, 3, 4, 9)]
    print(f"  首数据行 B/C/D/I = {r4}")

    # ── 2) 基本信息导出（整项目）──
    blob2 = fetch(f"{ADMIN}/api/projects/{PID}/export_base", cookie, timeout=300)
    print(f"基本信息导出: {len(blob2)} bytes")
    wb2 = openpyxl.load_workbook(io.BytesIO(blob2))
    results.append(("基本信息 3 sheet 生成", len(wb2.sheetnames) == 3))
    ws1 = wb2["2023年度人工造林"]
    # 找 B 列（调查小班号）= 1 的行（数据区 R5 起）
    row = None
    for r in range(5, 120):
        if ws1.cell(row=r, column=2).value == 1:
            row = r
            break
    if row is None:
        results.append(("找到小班1数据行", False))
    else:
        an = ws1.cell(row=row, column=40).value  # AN
        ao = ws1.cell(row=row, column=41).value  # AO
        ap = ws1.cell(row=row, column=42).value  # AP
        print(f"  小班1 行={row} AN={an!r} AO={ao!r} AP={ap!r}")
        results.append((f"AN 查数株数=860（B34新口径，旧口径Σ=43）（实际 {an!r}）", an == 860))
        results.append((f"AO 合格株树=820（实际 {ao!r}）", ao == 820))
        results.append((f"AP 合格率=95.35 数值（实际 {ap!r}）", ap == 95.35))

    print()
    all_ok = True
    for name, ok in results:
        print(f"{'PASS' if ok else 'FAIL'} {name}")
        all_ok = all_ok and ok
    sys.exit(0 if all_ok else 1)


if __name__ == "__main__":
    main()
