# -*- coding: utf-8 -*-
"""业务勾稽断言 — 证明导出结果满足 CONSTRAINTS.md 的业务不变量。

用法:
    python tools/check_invariants.py [--project-id xxx]   # 缺省取第一个项目

断言清单（对应约束编号）:
    I1 行数勾稽   基本信息每分类 sheet 数据行数 == DB 该分类小班数（B5）
    I2 排序勾稽   数据行按（林班, 调查小班号）数字序非降（D2）
    I3 株数勾稽   表1 AN=调查总株数(B34口径)、AP=Σ成活÷Σ种植（比率0-1）±0.0001（C9）
    I4 sheet 名   无 Excel 禁止字符且 ≤31 字符（D7）
    I5 块数勾稽   样地导出每 sheet 块数 == 该分类小班数（D6）
    I6 合格率类型 合格率列写数值（比率0-1）而非百分比字符串（D3）
任一 FAIL 以退出码 1 结束。
"""
import argparse
import re
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl  # noqa: E402

from survey.core import exporter, storage  # noqa: E402

_BAD_SHEET_CHARS = re.compile(r'[\[\]:*?/\\]')


class Ctx:
    """一次导出共享的上下文（导出文件只生成一次）。"""

    def __init__(self, pid):
        self.pid = pid
        self.tmp = tempfile.mkdtemp(prefix="hqz_invariants_")
        self.base_path, _ = exporter.export_base(pid, str(Path(self.tmp) / "base.xlsx"))
        self.samples_path, _ = exporter.export_samples(pid, str(Path(self.tmp) / "samples.xlsx"))
        self._base_wb = None
        self._samples_wb = None

    @property
    def base_wb(self):
        if self._base_wb is None:
            self._base_wb = openpyxl.load_workbook(self.base_path)
        return self._base_wb

    @property
    def samples_wb(self):
        if self._samples_wb is None:
            self._samples_wb = openpyxl.load_workbook(self.samples_path)
        return self._samples_wb

    def data_rows(self, ws):
        """基本信息 sheet 的数据行（B 列调查小班号非空，跳过表头 1-4 行）。"""
        return [r for r in range(5, ws.max_row + 1)
                if ws.cell(row=r, column=2).value not in (None, "")]


def _base_sheet(wb, cat):
    """基本信息导出中分类 cat 的 sheet（名动态年度：{N}年度{cat}，2026-08-25）。"""
    suffix = f"年度{cat}"
    for sn in wb.sheetnames:
        if sn.endswith(suffix):
            return sn
    return None


def i1_row_count(ctx):
    """I1 每分类 sheet 数据行数 == DB 该分类小班数"""
    problems = []
    for cat in exporter._TPL_SHEET_NAMES:
        sheet = _base_sheet(ctx.base_wb, cat)
        if not sheet:
            problems.append(f"[{cat}] 基本信息缺少 sheet（年度{cat}）")
            continue
        ws = ctx.base_wb[sheet]
        got = len(ctx.data_rows(ws))
        want = len(storage.list_project_subcompartment_rows(ctx.pid, category=cat))
        if got != want:
            problems.append(f"[{sheet}] 导出行数 {got} != DB 小班数 {want}")
    return problems


def i2_sort_order(ctx):
    """I2 数据行顺序 == DB 按（林班, 调查小班号）排序后的期望顺序（D2）

    注意 sheet 列布局不可直接当排序键：表1 F 列才是林班，表2 F 列是
    小班原始值——所以以 DB 侧排序结果为期望，比对小班号序列。
    """
    problems = []
    for cat in exporter._TPL_SHEET_NAMES:
        sheet = _base_sheet(ctx.base_wb, cat)
        if not sheet:
            continue
        rows = storage.list_project_subcompartment_rows(ctx.pid, category=cat)
        want = [exporter._num(r.get("subcompartment"))
                for r in sorted(rows, key=exporter._sc_sort_key)]
        ws = ctx.base_wb[sheet]
        got = [exporter._num(ws.cell(row=r, column=2).value) for r in ctx.data_rows(ws)]
        if got != want:
            bad = next((i for i, (g, w) in enumerate(zip(got, want)) if g != w),
                       min(len(got), len(want)))
            problems.append(f"[{sheet}] 第 {bad + 1} 个数据行起不符：导出 {got} != 期望 {want}")
    return problems


def i3_sample_stats(ctx):
    """I3 表1 AN=调查总株数（B34口径）、AP=Σ成活÷Σ种植 比率0-1 ±0.0001（有样地的小班）"""
    problems = []
    sheet = _base_sheet(ctx.base_wb, "人工造林")
    if not sheet:
        return ["基本信息缺少 sheet（年度人工造林）"]
    ws = ctx.base_wb[sheet]
    # DB 侧聚合：table1（人工造林）每小班 data（samples 存于各分类自身表 data_json，
    # 旧 table5 已随水利水保/草原下线删除——此前查 table5 恒空致 I3 假通过）
    db_stats = {}
    for rec in storage.get_survey_rows(ctx.pid, "table1"):
        st = exporter._sample_stats(rec.get("data") or {})
        if st["planted_total"]:
            db_stats[rec["subcompartment_id"]] = st
    # 导出侧：B 列小班号 → 行；需小班 id → 调查小班号映射
    sc_by_num = {}
    for r in storage.list_project_subcompartment_rows(ctx.pid, category="人工造林"):
        sc_by_num[exporter._num(r.get("subcompartment"))] = r["id"]
    for row in ctx.data_rows(ws):
        sc_num = exporter._num(ws.cell(row=row, column=2).value)
        st = db_stats.get(sc_by_num.get(sc_num))
        if not st:
            continue
        an = ws.cell(row=row, column=40).value   # AN 小班查数株数（调查总株数 B34 口径）
        ap = ws.cell(row=row, column=42).value   # AP 合格率（比率 0-1，0.00% 格式显示）
        if an != st["planted_total"]:
            problems.append(f"小班{sc_num} AN={an!r} != 调查总株数 {st['planted_total']}")
        if ap is None or abs(float(ap) - st["qualified_rate"]) > 0.0001:
            problems.append(f"小班{sc_num} AP={ap!r} != 合格率 {st['qualified_rate']}")
    return problems


def i4_sheet_names(ctx):
    """I4 sheet 名无禁止字符且 ≤31 字符"""
    problems = []
    for wb, tag in ((ctx.base_wb, "基本信息"), (ctx.samples_wb, "样地")):
        for sn in wb.sheetnames:
            if _BAD_SHEET_CHARS.search(sn) or len(sn) > 31:
                problems.append(f"{tag} sheet 名非法: {sn!r}")
    return problems


def i5_block_count(ctx):
    """I5 样地导出每 sheet 块数 == 该分类小班数"""
    problems = []
    cat_by_sheet = {v: k for k, v in exporter._TPL_SHEET_NAMES.items()}
    for sn in ctx.samples_wb.sheetnames:
        cat = cat_by_sheet.get(sn) or sn.split("-")[0]
        ws = ctx.samples_wb[sn]
        blocks = sum(1 for r in range(1, ws.max_row + 1, exporter._SAMPLE_BLOCK_ROWS)
                     if ws.cell(row=r, column=1).value not in (None, ""))
        want = len(storage.list_project_subcompartment_rows(ctx.pid, category=cat))
        if blocks != want:
            problems.append(f"[{sn}] 块数 {blocks} != DB 小班数 {want}")
    return problems


def i6_rate_is_number(ctx):
    """I6 合格率列写数值（int/float），非百分比字符串"""
    problems = []
    # 合格率列：表1 AP(42) / 表2 BG(59) / 表3 AT(46)（见 exporter._TPL_COL_MAPS）；
    # sheet 名年度动态（2026-08-25），按分类后缀查找
    for cat, col in (("人工造林", 42), ("封山育林", 59), ("退化林修复", 46)):
        sheet = _base_sheet(ctx.base_wb, cat)
        if not sheet:
            continue
        ws = ctx.base_wb[sheet]
        for r in ctx.data_rows(ws):
            v = ws.cell(row=r, column=col).value
            if v not in (None, "") and not isinstance(v, (int, float)):
                problems.append(f"[{sheet}] 行{r} 合格率 {v!r} 不是数值")
    return problems


CHECKS = [
    ("I1 行数勾稽", i1_row_count),
    ("I2 排序勾稽", i2_sort_order),
    ("I3 株数勾稽", i3_sample_stats),
    ("I4 sheet 名", i4_sheet_names),
    ("I5 块数勾稽", i5_block_count),
    ("I6 合格率类型", i6_rate_is_number),
]


def main():
    ap = argparse.ArgumentParser(description="导出业务勾稽断言")
    ap.add_argument("--project-id", help="项目 id（缺省取第一个项目）")
    args = ap.parse_args()

    pid = args.project_id or storage.list_projects()[0]["id"]
    name = storage.get_project(pid)["name"]
    print(f"项目: {name} ({pid})")
    ctx = Ctx(pid)

    failed = 0
    for label, fn in CHECKS:
        problems = fn(ctx)
        if problems:
            failed += 1
            print(f"FAIL {label}")
            for p in problems[:10]:
                print(f"     - {p}")
            if len(problems) > 10:
                print(f"     ... 其余 {len(problems) - 10} 条省略")
        else:
            print(f"PASS {label}")
    sys.exit(1 if failed else 0)


if __name__ == "__main__":
    main()
