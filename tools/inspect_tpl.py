# -*- coding: utf-8 -*-
"""Excel 模板结构探测 — 改模板或改导出代码前先跑，不猜结构（AGENTS.md）。

用法:
    python tools/inspect_tpl.py                 # 探测 tpl/ 下两个在用模板
    python tools/inspect_tpl.py <xlsx路径>      # 探测指定文件
    python tools/inspect_tpl.py --rows 8        # 每 sheet 打印前 8 行非空单元格

输出: sheet 名 / 维度 / 合并区 / 前 N 行非空单元格（值截断 20 字符）。
"""
import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl  # noqa: E402

from survey.core import exporter  # noqa: E402


def inspect(path, rows):
    print("=" * 70)
    print(f"文件: {path}")
    wb = openpyxl.load_workbook(path)
    print(f"sheet: {wb.sheetnames}")
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"\n── [{sn}] 维度 {ws.dimensions} (max_row={ws.max_row}, max_col={ws.max_column})")
        merged = list(ws.merged_cells.ranges)
        print(f"   合并区 {len(merged)} 个: {[str(m) for m in merged[:15]]}"
              + (" ..." if len(merged) > 15 else ""))
        for r in range(1, min(rows, ws.max_row) + 1):
            cells = []
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if v not in (None, ""):
                    s = str(v).replace("\n", "\\n")
                    cells.append(f"{ws.cell(row=r, column=c).coordinate}={s[:20]}")
            if cells:
                print(f"   R{r}: {' | '.join(cells[:12])}"
                      + (f" ... 共{len(cells)}格" if len(cells) > 12 else ""))
    wb.close()


def main():
    ap = argparse.ArgumentParser(description="Excel 模板结构探测")
    ap.add_argument("xlsx", nargs="?", help="xlsx 路径（缺省探测 tpl/ 两个在用模板）")
    ap.add_argument("--rows", type=int, default=6, help="每 sheet 打印前 N 行非空单元格")
    args = ap.parse_args()

    targets = [args.xlsx] if args.xlsx else [str(exporter._BASE_TEMPLATE), str(exporter._SAMPLE_TEMPLATE)]
    for t in targets:
        if not Path(t).exists():
            sys.exit(f"文件不存在: {t}")
        inspect(t, args.rows)


if __name__ == "__main__":
    main()
